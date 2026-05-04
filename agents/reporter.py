import json
import csv
import glob
from pathlib import Path
from typing import Dict, List
from collections import defaultdict
from rich.console import Console
from rich.table import Table
from rich import box
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from agents.io_utils import atomic_write_json

STATE_DIR = Path("state")
OUTPUT_DIR = Path("outputs")
WEB_DIR = Path("web")

DAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

LOCATION_FILES = {
    "Kwality House, Kemps Corner": "schedule_kwality.csv",
    "Supreme HQ, Bandra": "schedule_supreme.csv",
    "Kenkere House": "schedule_kenkere.csv",
}

LOCATION_EXCEL_FILES = {
    "Kwality House, Kemps Corner": "schedule_kwality.xlsx",
    "Supreme HQ, Bandra": "schedule_supreme.xlsx",
    "Kenkere House": "schedule_kenkere.xlsx",
}

CLASS_FAMILY_EXCEL_COLORS = {
    "barre":       ("E8D5FF", "6B21A8"),  # (bg, font) purple
    "powercycle":  ("FFD5D5", "991B1B"),  # red
    "strength_lab":("D5E4FF", "1E3A8A"),  # blue
    "recovery":    ("D5F5E8", "065F46"),  # green
    "foundations": ("FEF0D5", "92400E"),  # amber
    "mat_57":      ("D5F2F8", "0C4A6E"),  # cyan
    "hiit":        ("FFD5F0", "831843"),  # pink
    "default":     ("F0F0F5", "374151"),  # grey
}

FAMILY_COLORS = {
    "barre_57": "#7C3AED",
    "powercycle": "#DC2626",
    "strength_lab": "#1D4ED8",
    "recovery": "#059669",
    "foundations": "#D97706",
    "hiit": "#DB2777",
    "cardio": "#EA580C",
    "prenatal": "#65A30D",
    "special": "#6B7280",
}

OPTIMISATION_OPPORTUNITIES = [
    "Kenkere 10:00 AM — 56.5% fill, only 144 sessions. Expand to 5 days/week for highest ROI.",
    "Supreme 09:30 AM — 46.7% fill, 349 sessions. Add daily slot with Cauveri or Atulan.",
    "Anisha Shah at Kwality on Fridays — avg 7.7 check-in. 1 Friday morning slot would lift the weakest day.",
    "Kwality Thursday fill rate — 33.5%, lowest despite high volume. Swap 1-2 Thu slots to stronger class/trainer combos.",
    "Kenkere Saturday 4th morning — 40.6% fill. Kajol or Shruti Kulkarni as candidates.",
    "Supreme Friday underperformance — 31.2% fill. Test Vivaran or Cauveri on Fridays instead.",
]

ROOM_LABELS = {
  "studio_a": "Studio 1",
  "studio_b": "Studio 2",
  "powercycle": "PowerCycle Studio",
  "strength_lab": "Strength Lab",
}

MIN_SCHEDULE_SCORE = 80.0


def canonical_mix_class(class_name: str) -> str:
    lower = (class_name or "").lower()
    if "strength lab" in lower:
        return "Studio Strength Lab"
    if "back body blaze" in lower:
        return "Studio Back Body Blaze"
    if "powercycle" in lower or "power cycle" in lower:
        return "Studio PowerCycle"
    if "cardio barre" in lower:
        return "Studio Cardio Barre"
    if "mat 57" in lower:
        return "Studio Mat 57"
    if "barre 57" in lower:
        return "Studio Barre 57"
    if "amped" in lower:
        return "Studio Amped Up!"
    if "hiit" in lower:
        return "Studio HIIT"
    if "recovery" in lower:
        return "Studio Recovery"
    if "foundations" in lower:
        return "Studio Foundations"
    return class_name or "Unknown"


def _rules_panel_html() -> str:
    """Deprecated floating drawer — now replaced by the Rules nav tab in the main UI template."""
    return ""


class OutputReporter:
  def _display_room(self, room: str) -> str:
    return ROOM_LABELS.get(room, room)

  def _format_score_breakdown(self, breakdown: dict) -> str:
    if not breakdown:
        return ""
    parts = []
    for comp in breakdown.get("components", []):
        parts.append(
            f"{comp.get('label', comp.get('key', 'Component'))}: "
            f"{comp.get('points', 0)}/{comp.get('max_points', 0)}"
        )
    boost = breakdown.get("recency_boost", 0)
    parts.append(f"Recency: {boost:+.1f}")
    parts.append(f"Final: {breakdown.get('total_score', 0):.1f}")
    return " | ".join(parts)

  def _get_excel_family(self, class_name: str) -> str:
    lower = class_name.lower()
    if any(k in lower for k in ("barre 57", "cardio barre", "power barre", "barre fusion", "back body", "fit", "amped")):
        return "barre"
    if "powercycle" in lower or "power cycle" in lower:
        return "powercycle"
    if "strength" in lower:
        return "strength_lab"
    if "recovery" in lower or "flex" in lower:
        return "recovery"
    if "foundation" in lower:
        return "foundations"
    if "mat 57" in lower:
        return "mat_57"
    if "hiit" in lower or "dance cardio" in lower:
        return "hiit"
    return "default"

  def _iteration_names_for(self, all_schedules: list = None) -> list:
    defaults = ["Max Score", "Trainer Hours", "Class Variety"]
    if not all_schedules:
        return defaults
    names = []
    for idx, sched in enumerate(all_schedules[:3]):
        name = sched.get("iteration_name") if isinstance(sched, dict) else None
        names.append(name or defaults[idx])
    while len(names) < 3:
        names.append(defaults[len(names)])
    return names

  def _write_excel_multi_sheet(self, location: str, iterations_slots: list, week_start: str, sheet_names: list = None):
    """Write 3-sheet Excel file: Main, Iteration 2, Iteration 3."""
    from datetime import date, timedelta
    ws_date = date.fromisoformat(week_start)
    day_dates = {DAY_ORDER[i]: (ws_date + timedelta(days=i)).strftime("%-d-%b") for i in range(7)}

    filename = LOCATION_EXCEL_FILES.get(location, f"schedule_{location[:8].lower().replace(' ','_')}.xlsx")
    path = OUTPUT_DIR / filename

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    sheet_names = sheet_names or self._iteration_names_for()

    for iter_idx, (slots, sheet_name) in enumerate(zip(iterations_slots, sheet_names)):
        ws = wb.create_sheet(title=sheet_name)
        self._populate_excel_sheet(ws, location, slots, day_dates, week_start, sheet_name)

    wb.save(path)
    print(f"  [Excel] {path} written ({len(iterations_slots)} sheets)")

  def _populate_excel_sheet(self, ws, location: str, slots: list, day_dates: dict, week_start: str, label: str):
    """Populate one worksheet with a formatted schedule."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    time_fill = PatternFill("solid", fgColor="F5F5F8")
    time_font = Font(bold=True, color="555577", size=10)
    prime_fill = PatternFill("solid", fgColor="FFFBEB")
    prime_font = Font(bold=True, color="92400E", size=10)

    PRIME_TIMES = {"08:30","09:00","09:30","10:15","11:00","11:30","17:45","18:00","18:15","19:00","19:15","19:30"}

    # Row 1: location + label
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"{location}  |  {label}  |  Week of {week_start}"
    title_cell.font = Font(bold=True, size=13, color="1A1A2E")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    # Row 2: day headers (col A = Time, cols B-H = Mon-Sun)
    ws["A2"].value = "Time"
    ws["A2"].fill = header_fill
    ws["A2"].font = header_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].border = border

    for col_idx, day in enumerate(DAY_ORDER, start=2):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = f"{day}\n{day_dates[day]}"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 32

    # Build time -> day -> slot lookup
    by_time: dict = {}
    for slot in slots:
        t = slot["time"]
        day = slot["day_of_week"]
        if t not in by_time:
            by_time[t] = {}
        by_time[t][day] = slot

    all_times = sorted(by_time.keys())

    # Column widths
    ws.column_dimensions["A"].width = 7
    for col_idx in range(2, 9):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    for row_offset, t in enumerate(all_times):
        row = 3 + row_offset
        is_prime = t in PRIME_TIMES

        # Time column
        tc = ws.cell(row=row, column=1, value=t)
        tc.fill = prime_fill if is_prime else time_fill
        tc.font = prime_font if is_prime else time_font
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.border = border
        ws.row_dimensions[row].height = 52

        for col_idx, day in enumerate(DAY_ORDER, start=2):
            cell = ws.cell(row=row, column=col_idx)
            slot = by_time[t].get(day)
            if slot:
                family = self._get_excel_family(slot["class_name"])
                colors = CLASS_FAMILY_EXCEL_COLORS.get(family, CLASS_FAMILY_EXCEL_COLORS["default"])
                bg_color, font_color = colors
                cell.fill = PatternFill("solid", fgColor=bg_color)
                fill_pct = slot.get("predicted_fill_rate", 0)
                fill_str = f"{fill_pct:.0%}"
                hist_fill = slot.get("historical_avg_fill", 0)
                class_display = slot["class_name"].replace("Studio ", "")
                trainer = slot.get("trainer_1", "")
                rec = slot.get("recommendation", "")
                cell.value = f"{class_display}\n{trainer}\n{fill_str} pred | {hist_fill:.0%} hist\n[{rec}]"
                cell.font = Font(color=font_color, size=9)
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            else:
                cell.fill = PatternFill("solid", fgColor="FAFAFA")
            cell.border = border

    # Freeze panes: freeze row 2 and col A
    ws.freeze_panes = "B3"

  def run(self, all_schedules: list = None, primary_draft: dict = None) -> dict:
    print("[Agent 6] Reporter starting...")
    OUTPUT_DIR.mkdir(exist_ok=True)
    WEB_DIR.mkdir(exist_ok=True)

    # Determine primary draft to use for week_start and primary schedule
    if primary_draft is None and all_schedules:
        primary_draft = all_schedules[0]
    elif primary_draft is None:
        with open(STATE_DIR / "05_draft_schedule.json") as f:
            primary_draft = json.load(f)

    with open(STATE_DIR / "02_metrics.json") as f:
      metrics = json.load(f)
    score_baselines = self._build_score_baselines(primary_draft["schedule"])

    schedule = primary_draft["schedule"]
    week_start = primary_draft["target_week_start"]
    self._week_label = week_start
    iteration_names = self._iteration_names_for(all_schedules)

    by_location: Dict[str, List[dict]] = defaultdict(list)
    for slot in schedule:
      by_location[slot["location"]].append(slot)

    scorecard = {"generated_for_week": week_start, "locations": {}}

    # Build per-location data for all iterations
    all_by_location = []
    if all_schedules:
        for sched in all_schedules:
            loc_map: Dict[str, List[dict]] = defaultdict(list)
            for slot in sched["schedule"]:
                loc_map[slot["location"]].append(slot)
            all_by_location.append(loc_map)
    else:
        all_by_location = [by_location]

    # Use the selected schedule as Main for scorecard + CSV. When a composite
    # planner output includes iterations, iteration 0 may not be the chosen one.
    primary_by_location = by_location

    for loc, slots in primary_by_location.items():
      self._write_csv(loc, slots, week_start)
      self._write_detailed_csv(loc, slots, week_start)
      scorecard["locations"][loc] = self._build_scorecard_entry(loc, slots, score_baselines)

    # Write multi-sheet Excel per location
    all_locations = list(primary_by_location.keys())
    for loc in all_locations:
        iterations_slots = [loc_map.get(loc, []) for loc_map in all_by_location]
        # Pad to 3 if fewer iterations
        while len(iterations_slots) < 3:
            iterations_slots.append(iterations_slots[-1] if iterations_slots else [])
        self._write_excel_multi_sheet(loc, iterations_slots[:3], week_start, iteration_names)

    with open(OUTPUT_DIR / "scorecard.json", "w") as f:
      json.dump(scorecard, f, indent=2)

    # Generate web data and interface
    self._write_schedule_data(primary_by_location, week_start, metrics, all_by_location, iteration_names)
    self._generate_web_interface(primary_by_location, week_start, metrics, scorecard, all_by_location)

    self._print_summary(scorecard)
    self._run_assertions(scorecard, primary_by_location)
    self._run_iteration_score_assertions(all_by_location, all_schedules or [primary_draft], iteration_names)

    print(f"[Agent 6] Reporter complete — schedules + web UI written to {OUTPUT_DIR}/ and {WEB_DIR}/")
    return scorecard

    # ------------------------------------------------------------------ #
    #  CSV output
    # ------------------------------------------------------------------ #

  def _write_csv(self, location: str, slots: List[dict], week_start: str):
        from datetime import date, timedelta
        filename = LOCATION_FILES.get(location, f"schedule_{location[:8].lower().replace(' ','_')}.csv")
        path = OUTPUT_DIR / filename
        ws = date.fromisoformat(week_start)
        day_dates = {DAY_ORDER[i]: (ws + timedelta(days=i)).strftime("%-d-%b") for i in range(7)}

        by_time: Dict[str, Dict[str, dict]] = defaultdict(dict)
        for slot in slots:
            by_time[slot["time"]][slot["day_of_week"]] = slot

        all_times = sorted(by_time.keys())
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            # Enhanced header with detailed metrics
            writer.writerow([""] + [d for day in DAY_ORDER for d in [day_dates[day], "", "", "", "", "", "", "", "", ""]])
            writer.writerow([""] + [d for day in DAY_ORDER for d in [day, "", "", "", "", "", "", "", "", ""]])
            writer.writerow(["Time"] + [c for _ in DAY_ORDER for c in ["Class", "Trainer", "Room", "Cap", "Fill%", "Hist Sessions", "Hist Fill%", "Hist Avg", "Reason", ""]])
            for t in all_times:
                row = [t]
                for day in DAY_ORDER:
                    slot = by_time[t].get(day)
                    if slot:
                        fill_pct = f"{slot.get('predicted_fill_rate', 0):.0%}"
                        hist_sessions = slot.get("historical_session_count", 0)
                        hist_fill = f"{slot.get('historical_avg_fill', 0):.0%}"
                        hist_checkin = f"{slot.get('historical_avg_checkin', 0):.1f}"
                        reason = slot.get("scheduling_reason", "")[:30]  # Truncate for display
                        row += [
                            slot["class_name"],
                            slot["trainer_1"],
                            self._display_room(slot.get("room", "")),
                            slot.get("capacity", ""),
                            fill_pct,
                            hist_sessions,
                            hist_fill,
                            hist_checkin,
                            reason,
                            "",
                        ]
                    else:
                        row += ["", "", "", "", "", "", "", "", "", ""]
                writer.writerow(row)
            
            # Add breakdown sections for analysis
            writer.writerow([])
            writer.writerow([])
            writer.writerow(["=== WEEKLY BREAKDOWN ==="])
            writer.writerow([])
            
            # Trainer breakdown
            writer.writerow(["TRAINER ALLOCATION"])
            writer.writerow(["Trainer", "Total Classes", "Total Hours", "Avg Fill%", "Primary Classes"])
            trainer_stats = defaultdict(lambda: {"classes": 0, "hours": 0.0, "fill": [], "class_types": defaultdict(int)})
            for slot in slots:
                t = slot["trainer_1"]
                trainer_stats[t]["classes"] += 1
                trainer_stats[t]["hours"] += slot.get("duration_min", 57) / 60.0
                trainer_stats[t]["fill"].append(slot.get("predicted_fill_rate", 0))
                trainer_stats[t]["class_types"][slot["class_name"]] += 1
            
            for trainer in sorted(trainer_stats.keys()):
                stats = trainer_stats[trainer]
                avg_fill = sum(stats["fill"]) / len(stats["fill"]) if stats["fill"] else 0
                primary = max(stats["class_types"].items(), key=lambda x: x[1])[0] if stats["class_types"] else ""
                writer.writerow([
                    trainer,
                    stats["classes"],
                    f"{stats['hours']:.1f}",
                    f"{avg_fill:.0%}",
                    primary
                ])
            
            writer.writerow([])
            
            # Class format breakdown
            writer.writerow(["CLASS FORMAT ALLOCATION"])
            writer.writerow(["Class Format", "Total Classes", "Avg Fill%", "Days Scheduled"])
            format_stats = defaultdict(lambda: {"count": 0, "fill": [], "days": set()})
            for slot in slots:
                fmt = slot["class_name"]
                format_stats[fmt]["count"] += 1
                format_stats[fmt]["fill"].append(slot.get("predicted_fill_rate", 0))
                format_stats[fmt]["days"].add(slot["day_of_week"])
            
            for fmt in sorted(format_stats.keys()):
                stats = format_stats[fmt]
                avg_fill = sum(stats["fill"]) / len(stats["fill"]) if stats["fill"] else 0
                writer.writerow([
                    fmt,
                    stats["count"],
                    f"{avg_fill:.0%}",
                    len(stats["days"])
                ])
            
            writer.writerow([])
            
            # Timeslot breakdown
            writer.writerow(["TIMESLOT ALLOCATION"])
            writer.writerow(["Time", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun", "Total"])
            timeslot_grid = defaultdict(lambda: defaultdict(str))
            for slot in slots:
                timeslot_grid[slot["time"]][slot["day_of_week"]] = slot["class_name"]
            
            for time_slot in sorted(timeslot_grid.keys()):
                row = [time_slot]
                for day in DAY_ORDER:
                    row.append(timeslot_grid[time_slot].get(day, "-"))
                row.append(sum(1 for d in DAY_ORDER if timeslot_grid[time_slot].get(d)))
                writer.writerow(row)

    # ------------------------------------------------------------------ #
    #  Detailed wide CSV — one row per scheduled slot with full diagnostics
    # ------------------------------------------------------------------ #

  def _write_detailed_csv(self, location: str, slots: List[dict], week_start: str):
        suffix_map = {
            "Kwality House, Kemps Corner": "kwality",
            "Supreme HQ, Bandra": "supreme",
            "Kenkere House": "kenkere",
        }
        slug = suffix_map.get(location, location[:8].lower().replace(" ", "_"))
        path = OUTPUT_DIR / f"schedule_{slug}_detailed.csv"

        DAY_RANK = {d: i for i, d in enumerate(DAY_ORDER)}
        sorted_slots = sorted(
            slots,
            key=lambda s: (DAY_RANK.get(s.get("day_of_week", ""), 99), s.get("time", ""))
        )

        headers = [
            "Date", "Day", "Time", "Class", "Trainer 1", "Trainer 2", "Cover",
            "Room", "Capacity", "Duration (min)",
            "Score", "Recommendation", "Reason",
            "Score Breakdown",
            "Historic Sessions", "Historic Checkins", "Historic Fill Rate",
            "Projected Avg Checkin", "Projected Fill Rate",
            "Late Cancel Rate", "No Show Rate",
            "Constraint Violations",
        ]
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for s in sorted_slots:
                hist_fill = s.get("historical_avg_fill", 0) or 0
                hist_checkin = s.get("historical_avg_checkin", 0) or 0
                cap = s.get("capacity", 0) or 0
                projected_fill = s.get("predicted_fill_rate", 0) or 0
                projected_checkin = round(projected_fill * cap, 1) if cap else round(hist_checkin, 1)
                writer.writerow([
                    s.get("date", ""),
                    s.get("day_of_week", ""),
                    s.get("time", ""),
                    s.get("class_name", ""),
                    s.get("trainer_1", ""),
                    s.get("trainer_2", ""),
                    s.get("cover", ""),
                    self._display_room(s.get("room", "")),
                    cap,
                    s.get("duration_min", ""),
                    f"{s.get('score', 0):.1f}",
                    s.get("recommendation", ""),
                    s.get("scheduling_reason", ""),
                    self._format_score_breakdown(s.get("score_breakdown")),
                    s.get("historical_session_count", 0),
                    f"{hist_checkin:.2f}",
                    f"{hist_fill:.4f}",
                    projected_checkin,
                    f"{projected_fill:.4f}",
                    f"{s.get('historical_late_cancel_rate', 0):.4f}",
                    f"{s.get('historical_no_show_rate', 0):.4f}",
                    " | ".join(s.get("constraint_violations", []) or []),
                ])

    # ------------------------------------------------------------------ #
    #  Scorecard
    # ------------------------------------------------------------------ #

  def _build_score_baselines(self, schedule: List[dict]) -> Dict[str, float]:
        """Best-available average score for each location at the generated class count."""
        try:
            with open(STATE_DIR / "03_scores.json") as f:
                scores_data = json.load(f)
        except Exception:
            return {}

        counts: Dict[str, int] = defaultdict(int)
        for slot in schedule:
            counts[slot["location"]] += 1

        baselines: Dict[str, float] = {}
        for loc, count in counts.items():
            loc_scores = sorted(
                [
                    float(r.get("score", 0))
                    for r in scores_data.get("class_slot_ranking", [])
                    if r.get("location") == loc
                ],
                reverse=True,
            )
            if count and len(loc_scores) >= count:
                baselines[loc] = sum(loc_scores[:count]) / count
        return baselines

  def _calculate_schedule_score(self, location: str, slots: List[dict], score_baselines: Dict[str, float]) -> float:
        if not slots:
            return 0.0
        avg_slot_score = sum(float(s.get("score", 0)) for s in slots) / len(slots)
        baseline = score_baselines.get(location) or avg_slot_score or 1.0
        relative_performance = min(100.0, (avg_slot_score / max(baseline, 1.0)) * 100.0)
        violations = sum(len(s.get("constraint_violations", []) or []) for s in slots)
        compliance_score = max(0.0, 100.0 - violations * 10.0)
        return round(min(100.0, relative_performance * 0.75 + compliance_score * 0.25), 1)

  def _build_scorecard_entry(self, location: str, slots: List[dict], score_baselines: Dict[str, float] = None) -> dict:
        total = len(slots)
        if total == 0:
            return {}
        score_baselines = score_baselines or {}
        avg_fill = sum(s.get("predicted_fill_rate", 0) for s in slots) / total
        avg_slot_score = sum(float(s.get("score", 0)) for s in slots) / total
        schedule_score = self._calculate_schedule_score(location, slots, score_baselines)
        historical_avg_fill = sum(s.get("historical_avg_fill", 0) for s in slots) / total
        zero_history_slots = sum(1 for s in slots if s.get("historical_session_count", 0) == 0)
        class_counts: Dict[str, int] = defaultdict(int)
        for s in slots:
            class_counts[canonical_mix_class(s["class_name"])] += 1
        class_mix = {k: round(v / total, 3) for k, v in class_counts.items()}
        violations = [v for s in slots for v in s.get("constraint_violations", [])]
        exp_count = sum(1 for s in slots if s.get("is_experimental", False))

        # Barre family percentage for assertion checks
        barre_keywords = ["Barre", "FIT", "Mat 57", "Back Body", "Cardio", "Amped", "Power Barre", "Fusion"]
        barre_count = sum(1 for s in slots if any(kw in s["class_name"] for kw in barre_keywords))
        barre_pct = barre_count / total if total else 0

        # Format-specific counts for assertions
        format_counts = {
            "Studio Barre 57": class_counts.get("Studio Barre 57", 0),
            "Studio Mat 57": class_counts.get("Studio Mat 57", 0),
            "Studio FIT": class_counts.get("Studio FIT", 0),
            "Studio Cardio Barre": class_counts.get("Studio Cardio Barre", 0),
            "Studio PowerCycle": class_counts.get("Studio PowerCycle", 0),
            "Studio Strength Lab": class_counts.get("Studio Strength Lab", 0),
            "Studio HIIT": class_counts.get("Studio HIIT", 0),
            "Studio Amped Up!": class_counts.get("Studio Amped Up!", 0),
            "Studio Recovery": class_counts.get("Studio Recovery", 0),
            "Studio Back Body Blaze": class_counts.get("Studio Back Body Blaze", 0),
            "Studio Foundations": class_counts.get("Studio Foundations", 0),
        }

        # Strict Barre 57 family count (Barre 57, Cardio Barre, Power Barre, Barre Fusion, Barre 57 Express)
        barre_family_keywords = ["Barre 57", "Cardio Barre", "Power Barre", "Barre Fusion"]
        barre_family_count = sum(
            1 for s in slots if any(kw in s["class_name"] for kw in barre_family_keywords)
        )

        return {
            "total_classes": total,
            "target_schedule_score": int(MIN_SCHEDULE_SCORE),
            "schedule_score": schedule_score,
            "avg_slot_score": round(avg_slot_score, 1),
            "predicted_avg_fill_rate": round(avg_fill, 3),
            "historical_avg_fill_rate": round(historical_avg_fill, 3),
            "zero_history_slots": zero_history_slots,
            "zero_history_pct": round(zero_history_slots / total, 3) if total else 0,
            "experimental_pct": round(exp_count / total, 3) if total else 0,
            "barre_pct": round(barre_pct, 3),
            "barre_family_count": barre_family_count,
            "barre_family_pct": round(barre_family_count / total, 3) if total else 0,
            "format_counts": format_counts,
            "class_mix": class_mix,
            "hard_constraint_violations": list(set(violations)),
            "soft_constraint_penalties": 0,
            "optimisation_opportunities": OPTIMISATION_OPPORTUNITIES,
        }

  def _validate_against_schedule_config(self, by_location: Dict[str, List[dict]]) -> tuple[list[str], list[str]]:
        errors: list[str] = []
        warnings: list[str] = []
        config_path = Path("config/schedule_config.json")
        if not config_path.exists():
            return errors, warnings

        with open(config_path) as f:
            schedule_config = json.load(f)

        for loc, day_targets in schedule_config.get("targets", {}).items():
            loc_slots = by_location.get(loc, [])
            for day, limits in day_targets.items():
                if not isinstance(limits, dict):
                    continue
                target = limits.get("min", limits.get("target"))
                max_count = limits.get("max")
                actual = sum(1 for s in loc_slots if s.get("day_of_week") == day)
                if target is not None and actual < int(target):
                    errors.append(f"{loc} {day}: actual {actual} below min {int(target)}")
                if max_count is not None and actual > int(max_count):
                    errors.append(f"{loc} {day}: actual {actual} exceeds max {int(max_count)}")

        for loc, mix in schedule_config.get("class_mix", {}).items():
            loc_slots = by_location.get(loc, [])
            counts: Dict[str, int] = defaultdict(int)
            for slot in loc_slots:
                counts[canonical_mix_class(slot.get("class_name", ""))] += 1
            for class_name, limits in mix.items():
                if not isinstance(limits, dict):
                    continue
                actual = counts.get(class_name, 0)
                min_count = limits.get("min")
                max_count = limits.get("max")
                if min_count is not None and actual < int(min_count):
                    warnings.append(f"{loc}: {class_name} count {actual} < {int(min_count)}")
                if max_count is not None and actual > int(max_count):
                    warnings.append(f"{loc}: {class_name} count {actual} > {int(max_count)}")

        return errors, warnings

    # ------------------------------------------------------------------ #
    #  Web data JSON
    # ------------------------------------------------------------------ #

  def _write_schedule_data(self, by_location, week_start, metrics, all_by_location=None, iteration_names=None):
        from datetime import date, timedelta
        ws = date.fromisoformat(week_start)
        day_dates = {DAY_ORDER[i]: (ws + timedelta(days=i)).isoformat() for i in range(7)}

        # Build trainer metrics lookup
        trainer_metrics: Dict[str, dict] = {}
        for tm in metrics.get("trainer_metrics", []):
            key = f"{tm['location']}::{tm['trainer']}"
            trainer_metrics[key] = tm

        # Build scoring lookups from 03_scores.json for modal enrichment.
        # exact key: (location, class_name, trainer, day_of_week, time)
        # slot key:  (location, class_name, day_of_week, time)
        scores_lookup: Dict[tuple, dict] = {}
        slot_scores_lookup: Dict[tuple, dict] = {}
        scores_path = STATE_DIR / "03_scores.json"
        if scores_path.exists():
            try:
                with open(scores_path) as _sf:
                    _sd = json.load(_sf)
                for _sr in _sd.get("class_slot_ranking", []):
                    _k = (
                        _sr.get("location", ""),
                        _sr.get("class", ""),
                        _sr.get("trainer", ""),
                        _sr.get("day_name", ""),
                        _sr.get("time", ""),
                    )
                    scores_lookup[_k] = _sr
                for _sr in _sd.get("slot_group_ranking", []):
                    _k = (
                        _sr.get("location", ""),
                        _sr.get("class", ""),
                        _sr.get("day_name", ""),
                        _sr.get("time", ""),
                    )
                    slot_scores_lookup[_k] = _sr
            except Exception:
                pass

        web_data = {
            "generated_for_week": week_start,
            "day_dates": day_dates,
            "locations": {},
        }

        def _enrich_slot(s, loc, tm_map):
            trainer_key = f"{loc}::{s['trainer_1']}"
            tm = tm_map.get(trainer_key, {})
            sr_key = (
                loc,
                s.get("class_name", ""),
                s.get("trainer_1", ""),
                s.get("day_of_week", ""),
                s.get("time", ""),
            )
            slot_key = (
                loc,
                s.get("class_name", ""),
                s.get("day_of_week", ""),
                s.get("time", ""),
            )
            sr = scores_lookup.get(sr_key, {})
            slot_sr = slot_scores_lookup.get(slot_key, {})
            score_src = sr or slot_sr

            # Synthesize historic_detail from top-level fields when scorer didn't find exact match
            raw_hist = sr.get("historic_detail", None)
            raw_slot_hist = slot_sr.get("historic_detail", None)
            if not raw_hist and raw_slot_hist:
                raw_hist = {**raw_slot_hist, "_fallback_source": "day_class_time_location"}
            if raw_hist is None:
                avg_ci = s.get("historical_avg_checkin", 0) or 0
                avg_fill = s.get("historical_avg_fill", 0) or 0
                session_count = s.get("historical_session_count", 0) or 0
                if avg_ci > 0 or session_count > 0:
                    cap = s.get("capacity", 22) or 22
                    avg_booked = round(avg_ci / max(avg_fill, 0.05), 1) if avg_fill > 0 else avg_ci
                    raw_hist = {
                        "session_rows": session_count,
                        "avg_checked_in": avg_ci,
                        "avg_booked": avg_booked,
                        "avg_capacity": cap,
                        "avg_fill_rate": avg_fill,
                        "avg_revenue": 0,
                        "total_revenue": 0,
                        "avg_late_cancel_rate": s.get("historical_late_cancel_rate", 0) or 0,
                        "avg_no_show_rate": s.get("historical_no_show_rate", 0) or 0,
                        "individual_sessions": [],
                        "_synthetic": True,
                    }

            return {
                **s,
                "room": self._display_room(s.get("room", "")),
                "trainer_overall_fill": tm.get("trainer_fill_rate", 0),
                "trainer_overall_checkin": tm.get("trainer_avg_checkin", 0),
                "trainer_total_sessions": tm.get("trainer_session_count", 0),
                # Scoring breakdown fields (from scorer output)
                "blended_fill": score_src.get("blended_fill", None),
                "base_score": score_src.get("base_score", None),
                "recency_boost": score_src.get("recency_boost", None),
                "score_breakdown": score_src.get("score_breakdown", None),
                "historic_detail": raw_hist,
                "slot_historic_detail": raw_slot_hist,
                "slot_top_trainers": slot_sr.get("top_trainers", []),
                "slot_score_breakdown": slot_sr.get("score_breakdown", None),
                "slot_group_score": slot_sr.get("score", None),
                "slot_trust": score_src.get("slot_trust", None),
                "studio_avg_fill": score_src.get("studio_avg_fill", None),
                "above_studio_avg": score_src.get("above_studio_avg", None),
                "slot_sessions": score_src.get("session_count", None),
                "slot_avg_checkin": score_src.get("avg_checkin", score_src.get("avg_attendance", None)),
                "slot_avg_fill_rate": score_src.get("avg_fill_rate", None),
            }

        for loc, slots in by_location.items():
            web_data["locations"][loc] = [_enrich_slot(s, loc, trainer_metrics) for s in slots]

        # Add iteration data for web UI
        if all_by_location and len(all_by_location) > 1:
            iter_names = iteration_names or self._iteration_names_for()
            web_data["iteration_labels"] = {"Main": iter_names[0] if iter_names else "Max Score"}
            web_data["iterations"] = {}
            for iter_idx, loc_map in enumerate(all_by_location[1:3], start=1):
                iter_name = iter_names[iter_idx]
                web_data["iterations"][iter_name] = {}
                for loc, iter_slots in loc_map.items():
                    web_data["iterations"][iter_name][loc] = [
                        _enrich_slot(s, loc, trainer_metrics) for s in iter_slots
                    ]

        self._last_schedule_json = json.dumps(web_data, indent=2)
        atomic_write_json(WEB_DIR / "schedule_data.json", web_data, indent=2)
        versioned = WEB_DIR / f"schedule_data_{week_start}.json"
        atomic_write_json(versioned, web_data, indent=2)

    # ------------------------------------------------------------------ #
    #  Web HTML interface
    # ------------------------------------------------------------------ #

  def _generate_web_interface(self, by_location, week_start, metrics, scorecard, all_by_location=None):
        from datetime import date, timedelta
        ws = date.fromisoformat(week_start)
        week_end = ws + timedelta(days=6)
        week_label = f"{ws.strftime('%d %b')} – {week_end.strftime('%d %b %Y')}"

        # Use new template if available, fall back to inline generation
        template_path = WEB_DIR / "template.html"
        if template_path.exists():
            schedule_json = getattr(self, "_last_schedule_json", None)
            if schedule_json is None:
                with open(WEB_DIR / "schedule_data.json") as f:
                    schedule_json = f.read()
            template = template_path.read_text(encoding="utf-8")
            scorecard_json = json.dumps(scorecard, indent=2) if scorecard else "null"
            html = (template
                .replace("/*INJECT_SCHEDULE_DATA*/", schedule_json)
                .replace("/*INJECT_SCORECARD*/", scorecard_json)
                .replace("/*INJECT_WEEK_LABEL*/", f'"{week_label}"')
                .replace("/*INJECT_OPPORTUNITIES*/", json.dumps(OPTIMISATION_OPPORTUNITIES))
            )
            # Inject Rules panel into template path just before </body>
            html = html.replace("</body>", _rules_panel_html() + "\n</body>", 1)
            (WEB_DIR / "index.html").write_text(html, encoding="utf-8")
            print(f"[Agent 6] Web interface written to {WEB_DIR}/index.html (template-based)")
            print(f"[Agent 6] To view with rule toggles, run:")
            print(f"          python3 serve.py --week {getattr(self, '_week_label', week_start)} --port 8080")
            return

        schedule_json = getattr(self, "_last_schedule_json", None)
        if schedule_json is None:
            with open(WEB_DIR / "schedule_data.json") as f:
                schedule_json = f.read()

        optimisation_js = json.dumps(OPTIMISATION_OPPORTUNITIES)

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1.0" />
<title>Physique 57 — Schedule Intelligence</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
:root{{
  --bg:#FFFFFF;
  --surface:#F8FAFC;
  --surface2:#F1F5F9;
  --border:#E2E8F0;
  --border-strong:#CBD5E1;
  --text:#0F172A;
  --text-muted:#64748B;
  --text-light:#94A3B8;
  --primary:#1E40AF;
  --primary-light:#EFF6FF;
  --primary-mid:#DBEAFE;
  --accent:#F59E0B;
  --accent-light:#FFFBEB;
  --success:#16A34A;
  --danger:#DC2626;
  --warning:#D97706;
}}
html{{font-size:14px}}
body{{font-family:'Inter',sans-serif;background:var(--bg);color:var(--text);min-height:100vh;font-variant-numeric:tabular-nums}}

/* ---- SCROLLBAR ---- */
::-webkit-scrollbar{{width:5px;height:5px}}
::-webkit-scrollbar-track{{background:var(--surface)}}
::-webkit-scrollbar-thumb{{background:var(--border-strong);border-radius:3px}}

/* ---- HEADER ---- */
.app-header{{
  background:#FFFFFF;border-bottom:1px solid var(--border);
  padding:0 28px;height:56px;display:flex;align-items:center;
  justify-content:space-between;position:sticky;top:0;z-index:300;
}}
.header-left{{display:flex;align-items:center;gap:12px}}
.logo-mark{{
  width:32px;height:32px;background:var(--primary);border-radius:6px;
  display:flex;align-items:center;justify-content:center;
  font-size:11px;font-weight:800;color:#fff;letter-spacing:-0.5px;flex-shrink:0;
}}
.header-title{{font-size:15px;font-weight:700;color:var(--text)}}
.header-sub{{font-size:12px;color:var(--text-muted);font-weight:400}}
.week-pill{{
  background:var(--accent-light);color:var(--warning);
  border:1px solid #FDE68A;border-radius:100px;
  padding:4px 12px;font-size:12px;font-weight:600;
}}

/* ---- NAV BAR ---- */
.nav-bar{{
  background:#FFFFFF;border-bottom:1px solid var(--border);
  padding:0 28px;display:flex;align-items:center;
  justify-content:space-between;gap:16px;height:44px;
  position:sticky;top:56px;z-index:290;
}}
.nav-left{{display:flex;align-items:center;gap:2px}}
.nav-right{{display:flex;align-items:center;gap:8px}}
.loc-tab{{
  padding:6px 14px;border-radius:6px;font-size:13px;font-weight:600;
  cursor:pointer;border:none;background:transparent;color:var(--text-muted);
  transition:all 0.15s;white-space:nowrap;
}}
.loc-tab:hover{{background:var(--surface2);color:var(--text)}}
.loc-tab.active{{background:var(--primary-mid);color:var(--primary);}}
.view-btn{{
  padding:5px 11px;border-radius:6px;font-size:12px;font-weight:500;
  cursor:pointer;border:1px solid var(--border);background:var(--surface);
  color:var(--text-muted);transition:all 0.15s;display:flex;align-items:center;gap:5px;
}}
.view-btn:hover{{background:var(--surface2);color:var(--text)}}
.view-btn.active{{background:var(--primary);border-color:var(--primary);color:#fff;font-weight:600}}
.iter-pill{{
  padding:4px 10px;border-radius:100px;font-size:11px;font-weight:600;
  cursor:pointer;border:1px solid var(--border);background:var(--surface2);
  color:var(--text-muted);transition:all 0.15s;
}}
.iter-pill.active{{background:var(--accent);border-color:var(--accent);color:#fff}}
.nav-sep{{width:1px;height:20px;background:var(--border);margin:0 4px}}

/* ---- FILTER BAR ---- */
.filter-bar{{
  background:var(--surface);border-bottom:1px solid var(--border);
  padding:8px 28px;display:flex;align-items:center;gap:10px;flex-wrap:wrap;
  position:sticky;top:100px;z-index:280;
}}
.filter-label{{font-size:11px;font-weight:600;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.06em;white-space:nowrap}}
.filter-group{{display:flex;align-items:center;gap:4px;flex-wrap:wrap}}
.filter-sep{{width:1px;height:18px;background:var(--border-strong);}}
.day-toggle,.band-toggle,.rec-toggle{{
  padding:3px 9px;border-radius:100px;font-size:11px;font-weight:600;
  cursor:pointer;border:1px solid var(--border);background:#fff;
  color:var(--text-muted);transition:all 0.12s;user-select:none;
}}
.day-toggle.on{{background:var(--primary);border-color:var(--primary);color:#fff}}
.band-toggle.on{{background:var(--primary);border-color:var(--primary);color:#fff}}
.rec-toggle.on{{background:var(--primary-mid);border-color:var(--primary);color:var(--primary)}}
.filter-select{{
  padding:4px 8px;border:1px solid var(--border);border-radius:6px;
  font-size:12px;font-family:'Inter',sans-serif;background:#fff;
  color:var(--text);cursor:pointer;outline:none;
}}
.filter-clear{{
  padding:3px 10px;border-radius:100px;font-size:11px;font-weight:600;
  cursor:pointer;border:1px solid var(--border-strong);background:#fff;
  color:var(--text-muted);transition:all 0.12s;
}}
.filter-clear:hover{{background:var(--surface2)}}

/* ---- MAIN CONTENT ---- */
.main-area{{padding:20px 28px 40px}}

/* ---- STATS ROW ---- */
.stats-row{{display:flex;gap:10px;margin-bottom:18px;flex-wrap:wrap}}
.stat-card{{
  background:var(--surface);border:1px solid var(--border);border-radius:10px;
  padding:12px 16px;min-width:120px;flex:1;
}}
.stat-label{{font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:0.07em;color:var(--text-muted);margin-bottom:4px}}
.stat-value{{font-size:22px;font-weight:700;color:var(--text);line-height:1}}
.stat-value.primary{{color:var(--primary)}}
.stat-value.success{{color:var(--success)}}
.stat-value.warning{{color:var(--warning)}}
.stat-value.danger{{color:var(--danger)}}

/* ---- GRID VIEW ---- */
.schedule-grid{{
  display:grid;grid-template-columns:64px repeat(7,1fr);
  gap:2px;font-size:12px;
}}
.sg-corner,.sg-day-hdr{{
  background:var(--surface);border:1px solid var(--border);
  border-radius:6px;padding:8px;text-align:center;
  font-size:11px;font-weight:700;color:var(--text-muted);
  text-transform:uppercase;letter-spacing:0.05em;
}}
.sg-day-hdr{{color:var(--text)}}
.sg-day-hdr .date-sub{{font-size:10px;font-weight:400;color:var(--text-muted);margin-top:2px}}
.sg-time{{
  background:var(--surface);border:1px solid var(--border);border-radius:6px;
  padding:6px 8px;text-align:right;font-size:11px;font-weight:600;
  color:var(--text-muted);display:flex;align-items:flex-start;
  justify-content:flex-end;padding-top:9px;
}}
.sg-time.prime{{
  border-left:3px solid var(--accent);color:var(--warning);background:var(--accent-light);
}}
.sg-cell{{display:flex;flex-direction:column;gap:3px;min-height:36px}}

/* ---- CLASS CARDS ---- */
.cc{{
  border-radius:7px;padding:7px 9px;cursor:pointer;
  transition:box-shadow 0.15s,transform 0.1s;
  border-left:3px solid transparent;position:relative;
  overflow:hidden;
}}
.cc:hover{{transform:translateY(-1px);box-shadow:0 4px 12px rgba(0,0,0,0.1)}}
.cc-barre{{background:#EDE9FE;border-left-color:#7C3AED}}
.cc-powercycle{{background:#FEE2E2;border-left-color:#DC2626}}
.cc-strength_lab{{background:#DBEAFE;border-left-color:#1D4ED8}}
.cc-recovery{{background:#D1FAE5;border-left-color:#059669}}
.cc-foundations{{background:#FEF3C7;border-left-color:#D97706}}
.cc-mat_57{{background:#E0F2FE;border-left-color:#0284C7}}
.cc-hiit{{background:#FCE7F3;border-left-color:#DB2777}}
.cc-default{{background:#F1F5F9;border-left-color:#64748B}}
.cc-name{{font-weight:700;font-size:11px;line-height:1.25;color:var(--text);margin-bottom:2px}}
.cc-trainer{{font-size:10px;color:var(--text-muted);margin-bottom:3px}}
.cc-meta{{display:flex;align-items:center;gap:5px;flex-wrap:wrap}}
.cc-fill{{font-size:10px;font-weight:700}}
.cc-hist{{font-size:9px;color:var(--text-light)}}
.cc-badge{{
  font-size:8px;font-weight:700;padding:1px 5px;border-radius:100px;
  text-transform:uppercase;letter-spacing:0.05em;
}}
.cc-score{{font-size:9px;color:var(--text-light);margin-left:auto}}
.cc-exp{{
  position:absolute;top:3px;right:4px;
  font-size:8px;font-weight:700;color:var(--text-light);
  background:var(--surface2);padding:1px 4px;border-radius:3px;
}}
/* rec badge colors */
.cbadge-PINNED{{background:#D1FAE5;color:#065F46}}
.cbadge-PROTECT{{background:#EDE9FE;color:#4C1D95}}
.cbadge-INCLUDE{{background:#DBEAFE;color:#1E3A8A}}
.cbadge-CONSIDER{{background:#FEF3C7;color:#78350F}}
.cbadge-DROP{{background:#F1F5F9;color:#64748B}}

/* ---- LIST VIEW ---- */
.list-table{{width:100%;border-collapse:collapse;font-size:12px}}
.list-table thead th{{
  padding:8px 10px;text-align:left;font-size:10px;font-weight:700;
  text-transform:uppercase;letter-spacing:0.07em;color:var(--text-muted);
  background:var(--surface);border-bottom:2px solid var(--border);
  cursor:pointer;user-select:none;white-space:nowrap;position:sticky;top:0;
}}
.list-table thead th:hover{{color:var(--primary)}}
.list-table thead th.sort-asc::after{{content:" ▲";color:var(--primary)}}
.list-table thead th.sort-desc::after{{content:" ▼";color:var(--primary)}}
.list-table tbody tr{{border-bottom:1px solid var(--border);cursor:pointer;transition:background 0.1s}}
.list-table tbody tr:nth-child(even){{background:var(--surface)}}
.list-table tbody tr:hover{{background:var(--primary-light)}}
.list-table tbody tr.prime-row{{border-left:3px solid var(--accent)}}
.list-table td{{padding:7px 10px;color:var(--text);vertical-align:middle}}
.list-table td.td-muted{{color:var(--text-muted)}}
.viol-pill{{
  display:inline-flex;align-items:center;justify-content:center;
  background:#FEE2E2;color:#991B1B;border-radius:100px;
  font-size:10px;font-weight:700;width:20px;height:20px;
}}

/* ---- TRAINER VIEW ---- */
.trainer-section{{margin-bottom:20px}}
.trainer-hdr{{
  display:flex;align-items:center;gap:16px;
  background:var(--surface);border:1px solid var(--border);
  border-radius:10px;padding:12px 16px;margin-bottom:8px;
}}
.trainer-name{{font-size:14px;font-weight:700;color:var(--text)}}
.trainer-meta{{font-size:12px;color:var(--text-muted)}}
.trainer-mini-grid{{
  display:grid;grid-template-columns:80px repeat(7,1fr);
  gap:2px;font-size:11px;
}}
.tmg-lbl{{
  background:var(--surface);border:1px solid var(--border);
  border-radius:5px;padding:4px 6px;font-size:10px;
  font-weight:600;color:var(--text-muted);text-align:right;
}}
.tmg-day{{
  background:var(--surface);border:1px solid var(--border);
  border-radius:5px;padding:4px 6px;font-size:10px;
  font-weight:700;color:var(--text);text-align:center;
}}
.tmg-cell{{display:flex;flex-direction:column;gap:2px}}
.tmg-cls{{
  border-radius:4px;padding:3px 6px;font-size:10px;font-weight:600;
}}

/* ---- COMBINED VIEW ---- */
.combined-wrap{{overflow-x:auto}}
.combined-table{{min-width:1400px;border-collapse:collapse;font-size:11px}}
.combined-table th{{
  padding:6px 8px;background:var(--surface);border:1px solid var(--border);
  font-size:10px;font-weight:700;text-transform:uppercase;
  letter-spacing:0.05em;color:var(--text-muted);text-align:center;white-space:nowrap;
}}
.combined-table td{{
  padding:4px 6px;border:1px solid var(--border);vertical-align:top;
}}
.combined-table tr:nth-child(even) td{{background:var(--surface)}}
.ct-time{{font-size:10px;font-weight:700;color:var(--text-muted);white-space:nowrap;text-align:right;padding:4px 8px}}
.ct-time.prime{{color:var(--warning);background:var(--accent-light);border-left:3px solid var(--accent)}}
.ct-card{{border-radius:4px;padding:3px 5px;margin-bottom:2px;cursor:pointer}}
.ct-card:hover{{opacity:0.85}}
.ct-loc-kw{{font-size:9px;font-weight:700;color:var(--primary);text-align:center;padding:3px 0}}
.ct-loc-su{{font-size:9px;font-weight:700;color:#7C3AED;text-align:center;padding:3px 0}}
.ct-loc-ke{{font-size:9px;font-weight:700;color:var(--success);text-align:center;padding:3px 0}}

/* ---- ANALYTICS VIEW ---- */
.analytics-grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px}}
.analytics-card{{
  background:var(--surface);border:1px solid var(--border);
  border-radius:12px;padding:20px;
}}
.analytics-card.full{{grid-column:1/-1}}
.analytics-title{{font-size:13px;font-weight:700;color:var(--text);margin-bottom:14px}}
.analytics-sub{{font-size:11px;color:var(--text-muted);margin-bottom:8px}}
svg text{{font-family:'Inter',sans-serif}}
.opp-card{{
  background:#fff;border:1px solid var(--border);border-radius:8px;
  padding:12px 14px;margin-bottom:8px;border-left:3px solid var(--primary);
}}
.opp-card.red{{border-left-color:var(--danger)}}
.opp-text{{font-size:12px;color:var(--text);line-height:1.5}}
.trainer-util-row{{
  display:grid;grid-template-columns:140px 50px 50px 60px 1fr;
  gap:10px;align-items:center;padding:6px 0;
  border-bottom:1px solid var(--border);
}}
.util-bar-bg{{background:var(--surface2);border-radius:100px;height:8px;overflow:hidden}}
.util-bar-fill{{height:100%;border-radius:100px}}

/* ---- LEGEND ---- */
.legend-row{{display:flex;flex-wrap:wrap;gap:10px;margin-bottom:14px}}
.legend-item{{display:flex;align-items:center;gap:5px;font-size:11px;color:var(--text-muted)}}
.legend-dot{{width:10px;height:10px;border-radius:3px;flex-shrink:0}}

/* ---- MODAL ---- */
.modal-overlay{{
  position:fixed;inset:0;background:rgba(15,23,42,0.55);
  backdrop-filter:blur(4px);z-index:500;
  display:none;align-items:center;justify-content:center;
}}
.modal-overlay.open{{display:flex}}
.modal-box{{
  background:#fff;border:1px solid var(--border);border-radius:16px;
  width:580px;max-width:96vw;max-height:90vh;overflow-y:auto;
  box-shadow:0 20px 60px rgba(15,23,42,0.18);
  animation:modalIn 0.2s ease;
}}
@keyframes modalIn{{from{{transform:translateY(14px);opacity:0}}to{{transform:translateY(0);opacity:1}}}}
.modal-hdr{{
  padding:20px 24px 16px;border-bottom:1px solid var(--border);
  display:flex;align-items:flex-start;justify-content:space-between;
  gap:12px;
}}
.modal-close-btn{{
  width:32px;height:32px;border-radius:8px;border:1px solid var(--border);
  background:var(--surface);cursor:pointer;font-size:16px;
  display:flex;align-items:center;justify-content:center;flex-shrink:0;
  color:var(--text-muted);transition:all 0.12s;
}}
.modal-close-btn:hover{{background:var(--surface2);color:var(--text)}}
.modal-body{{padding:20px 24px}}
.modal-section{{margin-bottom:18px}}
.modal-section-ttl{{
  font-size:10px;font-weight:700;text-transform:uppercase;
  letter-spacing:0.1em;color:var(--text-muted);margin-bottom:10px;
}}
.modal-metric-grid{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px}}
.modal-metric{{
  background:var(--surface);border:1px solid var(--border);
  border-radius:8px;padding:10px 12px;
}}
.mm-lbl{{font-size:10px;color:var(--text-muted);margin-bottom:3px}}
.mm-val{{font-size:18px;font-weight:700;color:var(--text)}}
.modal-reason-box{{
  background:var(--primary-light);border-left:3px solid var(--primary);
  border-radius:0 8px 8px 0;padding:12px 14px;
  font-size:12px;line-height:1.6;color:var(--text);
}}
.chip-row{{display:flex;flex-wrap:wrap;gap:6px;margin-top:8px}}
.chip{{
  padding:4px 10px;border-radius:100px;font-size:11px;font-weight:600;
  border:1px solid;display:inline-block;
}}
.chip-prime{{background:var(--accent-light);border-color:#FDE68A;color:var(--warning)}}
.chip-exp{{background:var(--surface2);border-color:var(--border-strong);color:var(--text-muted)}}
.viol-tag{{
  display:inline-block;background:#FEE2E2;border:1px solid #FECACA;
  color:#991B1B;border-radius:5px;padding:3px 8px;font-size:11px;margin:2px;
}}
.prog-bar{{background:var(--surface2);border-radius:100px;height:6px;overflow:hidden;margin-top:4px}}
.prog-fill{{height:100%;border-radius:100px}}
.trainer-week-row{{display:flex;gap:6px;flex-wrap:wrap;margin-top:6px}}
.tw-chip{{
  background:var(--surface);border:1px solid var(--border);border-radius:5px;
  padding:4px 8px;font-size:10px;color:var(--text-muted);
}}
.tw-chip strong{{color:var(--text)}}
</style>
</head>
<body>

<!-- ============================================================
     HEADER
============================================================ -->
<header class="app-header">
  <div class="header-left">
    <div class="logo-mark">P57</div>
    <div>
      <span class="header-title">Schedule Intelligence</span>
      <span class="header-sub"> &nbsp;&middot;&nbsp; Physique 57 India</span>
    </div>
  </div>
  <div class="week-pill">Week of {week_label}</div>
</header>

<!-- ============================================================
     NAV BAR
============================================================ -->
<nav class="nav-bar">
  <div class="nav-left" id="loc-tabs">
    <!-- location tabs injected by JS -->
  </div>
  <div class="nav-right">
    <div id="iter-pills" style="display:flex;gap:4px"></div>
    <div class="nav-sep"></div>
    <button class="view-btn active" id="vbtn-grid"     onclick="setView('grid')"    >&#128197; Grid</button>
    <button class="view-btn"        id="vbtn-list"     onclick="setView('list')"    >&#9776; List</button>
    <button class="view-btn"        id="vbtn-trainer"  onclick="setView('trainer')" >&#128100; Trainer</button>
    <button class="view-btn"        id="vbtn-combined" onclick="setView('combined')">&#128279; Combined</button>
    <button class="view-btn"        id="vbtn-analytics"onclick="setView('analytics')">&#128202; Analytics</button>
  </div>
</nav>

<!-- ============================================================
     FILTER BAR
============================================================ -->
<div class="filter-bar" id="filter-bar">
  <span class="filter-label">Filter:</span>
  <div class="filter-group" id="day-toggles"></div>
  <div class="filter-sep"></div>
  <div class="filter-group" id="band-toggles">
    <span class="band-toggle on" data-band="morning"   onclick="toggleBand(this)">Morning</span>
    <span class="band-toggle on" data-band="midday"    onclick="toggleBand(this)">Midday</span>
    <span class="band-toggle on" data-band="evening"   onclick="toggleBand(this)">Evening</span>
  </div>
  <div class="filter-sep"></div>
  <select class="filter-select" id="class-filter" onchange="applyFilters()">
    <option value="">All Classes</option>
  </select>
  <select class="filter-select" id="trainer-filter" onchange="applyFilters()">
    <option value="">All Trainers</option>
  </select>
  <div class="filter-group" id="rec-toggles">
    <span class="rec-toggle on" data-rec="PINNED"  onclick="toggleRec(this)">PINNED</span>
    <span class="rec-toggle on" data-rec="PROTECT" onclick="toggleRec(this)">PROTECT</span>
    <span class="rec-toggle on" data-rec="INCLUDE" onclick="toggleRec(this)">INCLUDE</span>
    <span class="rec-toggle on" data-rec="CONSIDER"onclick="toggleRec(this)">CONSIDER</span>
    <span class="rec-toggle on" data-rec="DROP"    onclick="toggleRec(this)">DROP</span>
  </div>
  <button class="filter-clear" onclick="clearFilters()">&#10005; Clear</button>
</div>

<!-- ============================================================
     MAIN AREA
============================================================ -->
<div class="main-area" id="main-area"></div>

<!-- ============================================================
     MODAL
============================================================ -->
<div class="modal-overlay" id="modal-overlay">
  <div class="modal-box" id="modal-box"></div>
</div>

<script>
// ============================================================
// DATA
// ============================================================
const SCHEDULE_DATA = {schedule_json};
const OPTIMISATION_OPPORTUNITIES = {optimisation_js};

const DAY_ORDER = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"];
const DAY_SHORT  = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"];

const PRIME_TIMES = new Set(["08:30","09:00","09:30","10:15","11:00","11:30","17:45","18:00","18:15","19:00","19:15","19:30"]);

const ROOM_LABELS = {{
  "studio_a":"Studio 1","studio_b":"Studio 2",
  "powercycle":"PowerCycle Studio","strength_lab":"Strength Lab"
}};

const FAMILY_CSS = {{
  barre:"cc-barre", powercycle:"cc-powercycle", strength_lab:"cc-strength_lab",
  recovery:"cc-recovery", foundations:"cc-foundations", mat_57:"cc-mat_57",
  hiit:"cc-hiit", default:"cc-default"
}};
const FAMILY_BG = {{
  barre:"#EDE9FE", powercycle:"#FEE2E2", strength_lab:"#DBEAFE",
  recovery:"#D1FAE5", foundations:"#FEF3C7", mat_57:"#E0F2FE",
  hiit:"#FCE7F3", default:"#F1F5F9"
}};
const FAMILY_BORDER = {{
  barre:"#7C3AED", powercycle:"#DC2626", strength_lab:"#1D4ED8",
  recovery:"#059669", foundations:"#D97706", mat_57:"#0284C7",
  hiit:"#DB2777", default:"#64748B"
}};
const FAMILY_TEXT = {{
  barre:"#4C1D95", powercycle:"#7F1D1D", strength_lab:"#1E3A8A",
  recovery:"#064E3B", foundations:"#78350F", mat_57:"#0C4A6E",
  hiit:"#831843", default:"#1E293B"
}};

// ============================================================
// HELPERS
// ============================================================
function getFamily(cn) {{
  if (!cn) return "default";
  const l = cn.toLowerCase();
  if (l.includes("powercycle")||l.includes("power cycle")) return "powercycle";
  if (l.includes("strength lab")) return "strength_lab";
  if (l.includes("recovery")||l.includes("flex & flow")) return "recovery";
  if (l.includes("foundation")) return "foundations";
  if (l.includes("mat 57")) return "mat_57";
  if (l.includes("hiit")||l.includes("dance cardio")) return "hiit";
  if (l.includes("barre")||l.includes("fit")||l.includes("amped")||l.includes("back body")) return "barre";
  return "default";
}}

function tmins(t) {{
  if (!t) return 0;
  const [h,m] = t.split(":").map(Number); return h*60+m;
}}

function isPrime(t) {{ return PRIME_TIMES.has(t); }}

function pct(v,decimals=0) {{
  if (v==null||isNaN(v)) return "—";
  return (v*100).toFixed(decimals)+"%";
}}

function fillClr(v) {{
  if (v==null) return "#94A3B8";
  if (v>=0.6) return "#16A34A";
  if (v>=0.4) return "#D97706";
  return "#DC2626";
}}

function timeBand(t) {{
  const m=tmins(t);
  if (m>=7*60 && m<10*60) return "morning";
  if (m>=10*60 && m<13*60) return "midday";
  return "evening";
}}

function displayRoom(r) {{
  return ROOM_LABELS[r]||r||"";
}}

function displayClass(cn) {{
  return (cn||"").replace(/^Studio /,"");
}}

function h(tag,attrs,children) {{
  const el=document.createElement(tag);
  if (attrs) Object.entries(attrs).forEach(([k,v])=>{{
    if (k==="className") el.className=v;
    else if (k==="onclick") el.onclick=v;
    else if (k==="innerHTML") el.innerHTML=v;
    else if (k==="style") el.style.cssText=v;
    else if (k==="title") el.title=v;
    else el.setAttribute(k,v);
  }});
  if (children) {{
    (Array.isArray(children)?children:[children]).forEach(c=>{{
      if (c==null) return;
      el.appendChild(typeof c==="string"?document.createTextNode(c):c);
    }});
  }}
  return el;
}}

// ============================================================
// STATE
// ============================================================
let _loc = null;
let _view = "grid";
let _iter = "Main";
let _activeDays = new Set(DAY_ORDER);
let _activeBands = new Set(["morning","midday","evening"]);
let _activeRecs = new Set(["PINNED","PROTECT","INCLUDE","CONSIDER","DROP"]);
let _classFilter = "";
let _trainerFilter = "";
let _sortCol = null;
let _sortDir = 1;

function getSlots(loc, iter) {{
  loc = loc || _loc;
  iter = iter || _iter;
  if (iter !== "Main" && SCHEDULE_DATA.iterations && SCHEDULE_DATA.iterations[iter] && SCHEDULE_DATA.iterations[iter][loc]) {{
    return SCHEDULE_DATA.iterations[iter][loc] || [];
  }}
  return (SCHEDULE_DATA.locations && SCHEDULE_DATA.locations[loc]) || [];
}}

function filterSlots(slots) {{
  return slots.filter(s => {{
    if (!_activeDays.has(s.day_of_week)) return false;
    if (!_activeBands.has(timeBand(s.time))) return false;
    const rec = s.recommendation || "INCLUDE";
    if (!_activeRecs.has(rec)) return false;
    if (_classFilter && !s.class_name.toLowerCase().includes(_classFilter.toLowerCase())) return false;
    if (_trainerFilter && (s.trainer_1||"").toLowerCase()!==_trainerFilter.toLowerCase()) return false;
    return true;
  }});
}}

// ============================================================
// INIT
// ============================================================
window.addEventListener("DOMContentLoaded", () => {{
  buildLocTabs();
  buildDayToggles();
  buildIterPills();
  const locs = Object.keys(SCHEDULE_DATA.locations||{{}});
  if (locs.length) {{
    _loc = locs[0];
    document.getElementById("loc-tabs").children[0].classList.add("active");
    populateFilterDropdowns(getSlots(_loc));
    renderView();
  }}
}});

function buildLocTabs() {{
  const locs = Object.keys(SCHEDULE_DATA.locations||{{}});
  const bar = document.getElementById("loc-tabs");
  locs.forEach(loc => {{
    const btn = h("button", {{
      className:"loc-tab",
      title:loc,
      onclick:()=>switchLoc(loc, btn)
    }}, loc.split(",")[0]);
    bar.appendChild(btn);
  }});
}}

function buildDayToggles() {{
  const bar = document.getElementById("day-toggles");
  DAY_ORDER.forEach((day,i) => {{
    const el = h("span", {{
      className:"day-toggle on",
      "data-day":day,
      onclick:()=>toggleDay(el)
    }}, DAY_SHORT[i]);
    bar.appendChild(el);
  }});
}}

function buildIterPills() {{
  const bar = document.getElementById("iter-pills");
  const iters = ["Main"];
  if (SCHEDULE_DATA.iterations) {{
    Object.keys(SCHEDULE_DATA.iterations).forEach(k=>{{ if(k!=="Main") iters.push(k); }});
  }}
  if (iters.length <= 1) return;
  iters.forEach(iter => {{
    const btn = h("button", {{
      className:"iter-pill" + (iter==="Main"?" active":""),
      onclick:()=>switchIter(iter, btn)
    }}, iter);
    bar.appendChild(btn);
  }});
}}

function populateFilterDropdowns(slots) {{
  const classes = [...new Set(slots.map(s=>s.class_name).filter(Boolean))].sort();
  const trainers = [...new Set(slots.map(s=>s.trainer_1).filter(Boolean))].sort();
  const cf = document.getElementById("class-filter");
  const tf = document.getElementById("trainer-filter");
  cf.innerHTML = '<option value="">All Classes</option>';
  tf.innerHTML = '<option value="">All Trainers</option>';
  classes.forEach(c=>{{ const o=document.createElement("option"); o.value=c; o.textContent=displayClass(c); cf.appendChild(o); }});
  trainers.forEach(t=>{{ const o=document.createElement("option"); o.value=t; o.textContent=t; tf.appendChild(o); }});
}}

function switchLoc(loc, btn) {{
  _loc = loc;
  document.querySelectorAll(".loc-tab").forEach(b=>b.classList.remove("active"));
  btn.classList.add("active");
  populateFilterDropdowns(getSlots(loc));
  renderView();
}}

function switchIter(iter, btn) {{
  _iter = iter;
  document.querySelectorAll(".iter-pill").forEach(b=>b.classList.remove("active"));
  btn.classList.add("active");
  renderView();
}}

function setView(v) {{
  _view = v;
  document.querySelectorAll(".view-btn").forEach(b=>b.classList.remove("active"));
  const el = document.getElementById("vbtn-"+v);
  if (el) el.classList.add("active");
  renderView();
}}

// ============================================================
// FILTERS
// ============================================================
function toggleDay(el) {{
  const day = el.dataset.day;
  el.classList.toggle("on");
  if (el.classList.contains("on")) _activeDays.add(day);
  else _activeDays.delete(day);
  renderView();
}}

function toggleBand(el) {{
  const band = el.dataset.band;
  el.classList.toggle("on");
  if (el.classList.contains("on")) _activeBands.add(band);
  else _activeBands.delete(band);
  renderView();
}}

function toggleRec(el) {{
  const rec = el.dataset.rec;
  el.classList.toggle("on");
  if (el.classList.contains("on")) _activeRecs.add(rec);
  else _activeRecs.delete(rec);
  renderView();
}}

function applyFilters() {{
  _classFilter   = document.getElementById("class-filter").value;
  _trainerFilter = document.getElementById("trainer-filter").value;
  renderView();
}}

function clearFilters() {{
  _activeDays = new Set(DAY_ORDER);
  _activeBands = new Set(["morning","midday","evening"]);
  _activeRecs = new Set(["PINNED","PROTECT","INCLUDE","CONSIDER","DROP"]);
  _classFilter = "";
  _trainerFilter = "";
  document.querySelectorAll(".day-toggle,.band-toggle,.rec-toggle").forEach(el=>el.classList.add("on"));
  document.getElementById("class-filter").value = "";
  document.getElementById("trainer-filter").value = "";
  renderView();
}}

// ============================================================
// ROUTER
// ============================================================
function renderView() {{
  const area = document.getElementById("main-area");
  area.innerHTML = "";
  if (!_loc) return;
  const slots = getSlots();
  const filtered = filterSlots(slots);
  if (_view === "grid")      renderGrid(area, filtered, slots);
  else if (_view === "list") renderList(area, filtered);
  else if (_view === "trainer") renderTrainer(area, filtered);
  else if (_view === "combined") renderCombined(area);
  else if (_view === "analytics") renderAnalytics(area, slots);
}}

// ============================================================
// STATS ROW
// ============================================================
function makeStatsRow(slots) {{
  const total = slots.length;
  const avgFill = total ? slots.reduce((a,s)=>a+(s.predicted_fill_rate||0),0)/total : 0;
  const protect = slots.filter(s=>s.recommendation==="PROTECT"||s.recommendation==="PINNED").length;
  const viols   = slots.filter(s=>(s.constraint_violations||[]).length>0).length;
  const div = h("div",{{className:"stats-row"}},[
    statCard("Total Classes",total,"primary"),
    statCard("Avg Predicted Fill",pct(avgFill),"success"),
    statCard("Pinned/Protected",protect,"warning"),
    statCard("Violation Flags",viols, viols>0?"danger":"primary"),
    statCard("Prime Slots",slots.filter(s=>isPrime(s.time)).length,"primary"),
  ]);
  return div;
}}

function statCard(label, value, cls) {{
  return h("div",{{className:"stat-card"}},[
    h("div",{{className:"stat-label"}},label),
    h("div",{{className:`stat-value ${{cls}}`}},String(value)),
  ]);
}}

// ============================================================
// CLASS CARD ELEMENT
// ============================================================
function makeClassCard(s) {{
  const fam = getFamily(s.class_name);
  const card = h("div",{{className:`cc ${{FAMILY_CSS[fam]||"cc-default"}}`,onclick:()=>openModal(s)}});
  if (s.is_experimental) card.appendChild(h("span",{{className:"cc-exp"}},"EXP"));
  card.appendChild(h("div",{{className:"cc-name"}},displayClass(s.class_name)));
  card.appendChild(h("div",{{className:"cc-trainer"}},s.trainer_1||"—"));
  const meta = h("div",{{className:"cc-meta"}});
  const fillEl = h("span",{{className:"cc-fill",style:`color:${{fillClr(s.predicted_fill_rate)}}`}},pct(s.predicted_fill_rate));
  const histEl = h("span",{{className:"cc-hist"}},`hist ${{pct(s.historical_avg_fill)}}`);
  const rec = s.recommendation||"INCLUDE";
  const badge = h("span",{{className:`cc-badge cbadge-${{rec}}`}},rec);
  const score = h("span",{{className:"cc-score"}},`${{(s.score||0).toFixed(0)}}`);
  meta.appendChild(fillEl);
  meta.appendChild(histEl);
  meta.appendChild(badge);
  meta.appendChild(score);
  card.appendChild(meta);
  return card;
}}

// ============================================================
// VIEW 1: GRID
// ============================================================
function renderGrid(area, slots, allSlots) {{
  area.appendChild(makeStatsRow(allSlots));

  // legend
  const legRow = h("div",{{className:"legend-row"}});
  [["Barre Family","#7C3AED"],["PowerCycle","#DC2626"],["Strength Lab","#1D4ED8"],
   ["Recovery","#059669"],["Foundations","#D97706"],["Mat 57","#0284C7"],
   ["HIIT","#DB2777"],["Other","#64748B"]].forEach(([lbl,clr])=>{{
    legRow.appendChild(h("div",{{className:"legend-item"}},[
      h("div",{{className:"legend-dot",style:`background:${{clr}}`}}),
      lbl
    ]));
  }});
  area.appendChild(legRow);

  const byTime = {{}};
  slots.forEach(s=>{{
    if (!byTime[s.time]) byTime[s.time]={{}};
    if (!byTime[s.time][s.day_of_week]) byTime[s.time][s.day_of_week]=[];
    byTime[s.time][s.day_of_week].push(s);
  }});
  const times = Object.keys(byTime).sort();

  const grid = h("div",{{className:"schedule-grid"}});
  grid.appendChild(h("div",{{className:"sg-corner"}}));
  DAY_ORDER.forEach(day=>{{
    const dd = SCHEDULE_DATA.day_dates ? SCHEDULE_DATA.day_dates[day] : "";
    const dateStr = dd ? new Date(dd).toLocaleDateString("en-IN",{{day:"2-digit",month:"short"}}) : "";
    const cel = h("div",{{className:"sg-day-hdr"}});
    cel.innerHTML = `<div>${{day.slice(0,3)}}</div><div class="date-sub">${{dateStr}}</div>`;
    grid.appendChild(cel);
  }});

  times.forEach(t=>{{
    const prime = isPrime(t);
    grid.appendChild(h("div",{{className:"sg-time"+(prime?" prime":"")}},t));
    DAY_ORDER.forEach(day=>{{
      const cell = h("div",{{className:"sg-cell"}});
      const daySlots = (byTime[t]&&byTime[t][day])||[];
      daySlots.forEach(s=>cell.appendChild(makeClassCard(s)));
      grid.appendChild(cell);
    }});
  }});
  area.appendChild(grid);
}}

// ============================================================
// VIEW 2: LIST
// ============================================================
const LIST_COLS = [
  {{key:"day_of_week",label:"Day",sort:s=>DAY_ORDER.indexOf(s.day_of_week)}},
  {{key:"time",label:"Time",sort:s=>tmins(s.time)}},
  {{key:"_prime",label:"Prime",sort:s=>isPrime(s.time)?1:0}},
  {{key:"class_name",label:"Class",sort:s=>s.class_name||""}},
  {{key:"trainer_1",label:"Trainer",sort:s=>s.trainer_1||""}},
  {{key:"room",label:"Room",sort:s=>s.room||""}},
  {{key:"predicted_fill_rate",label:"Fill%",sort:s=>s.predicted_fill_rate||0}},
  {{key:"historical_avg_fill",label:"Hist Fill%",sort:s=>s.historical_avg_fill||0}},
  {{key:"historical_session_count",label:"Hist Sessions",sort:s=>s.historical_session_count||0}},
  {{key:"score",label:"Score",sort:s=>s.score||0}},
  {{key:"recommendation",label:"Rec",sort:s=>s.recommendation||""}},
  {{key:"_violations",label:"Viols",sort:s=>(s.constraint_violations||[]).length}},
];

function renderList(area, slots) {{
  // sort
  let sorted = [...slots];
  if (_sortCol!=null) {{
    const col = LIST_COLS[_sortCol];
    sorted.sort((a,b)=>{{
      const av=col.sort(a), bv=col.sort(b);
      if (av<bv) return -_sortDir;
      if (av>bv) return _sortDir;
      return 0;
    }});
  }} else {{
    sorted.sort((a,b)=>DAY_ORDER.indexOf(a.day_of_week)-DAY_ORDER.indexOf(b.day_of_week)||tmins(a.time)-tmins(b.time));
  }}

  const wrap = h("div",{{}});
  const table = h("table",{{className:"list-table"}});
  const thead = h("thead");
  const hrow = h("tr");
  LIST_COLS.forEach((col,i)=>{{
    const th = h("th",{{onclick:()=>sortList(i)}},col.label);
    if (_sortCol===i) th.className = _sortDir===1?"sort-asc":"sort-desc";
    hrow.appendChild(th);
  }});
  thead.appendChild(hrow);
  table.appendChild(thead);

  const tbody = h("tbody");
  sorted.forEach(s=>{{
    const prime = isPrime(s.time);
    const tr = h("tr",{{
      className:prime?"prime-row":"",
      onclick:()=>openModal(s)
    }});
    const viols = (s.constraint_violations||[]).length;
    const vals = [
      s.day_of_week,
      s.time,
      prime?"★":"",
      displayClass(s.class_name),
      s.trainer_1||"—",
      displayRoom(s.room),
      pct(s.predicted_fill_rate),
      pct(s.historical_avg_fill),
      String(s.historical_session_count||0),
      (s.score||0).toFixed(1),
      s.recommendation||"INCLUDE",
      "",
    ];
    vals.forEach((v,ci)=>{{
      const td = h("td",{{}});
      if (ci===2 && prime) td.style.cssText="color:var(--accent);font-weight:700";
      else if (ci===6) td.style.cssText=`color:${{fillClr(s.predicted_fill_rate)}};font-weight:700`;
      else if (ci===7) td.style.cssText=`color:${{fillClr(s.historical_avg_fill)}}`;
      if (ci===11) {{
        if (viols>0) td.appendChild(h("span",{{className:"viol-pill"}},String(viols)));
      }} else {{
        td.textContent = v;
      }}
      tr.appendChild(td);
    }});
    tbody.appendChild(tr);
  }});
  table.appendChild(tbody);
  wrap.appendChild(table);
  area.appendChild(wrap);
}}

function sortList(colIdx) {{
  if (_sortCol===colIdx) _sortDir*=-1;
  else {{ _sortCol=colIdx; _sortDir=1; }}
  renderView();
}}

// ============================================================
// VIEW 3: TRAINER
// ============================================================
function renderTrainer(area, slots) {{
  const byTrainer = {{}};
  slots.forEach(s=>{{
    const t = s.trainer_1||"Unknown";
    if (!byTrainer[t]) byTrainer[t]=[];
    byTrainer[t].push(s);
  }});

  const sorted = Object.entries(byTrainer).sort((a,b)=>b[1].length-a[1].length||a[0].localeCompare(b[0]));

  sorted.forEach(([trainer, tslots])=>{{
    const avgFill = tslots.reduce((a,s)=>a+(s.predicted_fill_rate||0),0)/tslots.length;
    const hours = tslots.reduce((a,s)=>a+(s.duration_min||57)/60,0);

    const sec = h("div",{{className:"trainer-section"}});
    const hdr = h("div",{{className:"trainer-hdr"}});
    hdr.innerHTML = `
      <div class="trainer-name">${{trainer}}</div>
      <div class="trainer-meta">Classes this week: <strong>${{tslots.length}}</strong></div>
      <div class="trainer-meta">Total hours: <strong>${{hours.toFixed(1)}}</strong></div>
      <div class="trainer-meta">Avg fill: <strong style="color:${{fillClr(avgFill)}}">${{pct(avgFill)}}</strong></div>
    `;
    sec.appendChild(hdr);

    // mini grid
    const grid = h("div",{{className:"trainer-mini-grid"}});
    grid.appendChild(h("div",{{className:"tmg-lbl"}},"Time"));
    DAY_ORDER.forEach((d,i)=>grid.appendChild(h("div",{{className:"tmg-day"}},DAY_SHORT[i])));

    const byTime2={{}};
    tslots.forEach(s=>{{
      if (!byTime2[s.time]) byTime2[s.time]={{}};
      byTime2[s.time][s.day_of_week]=s;
    }});
    const times2 = Object.keys(byTime2).sort();
    times2.forEach(t=>{{
      grid.appendChild(h("div",{{className:"tmg-lbl",style:isPrime(t)?"color:var(--warning)":""}},t));
      DAY_ORDER.forEach(day=>{{
        const cell = h("div",{{className:"tmg-cell"}});
        const s = byTime2[t][day];
        if (s) {{
          const fam=getFamily(s.class_name);
          const cls=h("div",{{
            className:"tmg-cls",
            style:`background:${{FAMILY_BG[fam]}};color:${{FAMILY_TEXT[fam]}};cursor:pointer`,
            onclick:()=>openModal(s),
            title:s.class_name
          }},displayClass(s.class_name).slice(0,14));
          cell.appendChild(cls);
        }}
        grid.appendChild(cell);
      }});
    }});
    sec.appendChild(grid);
    area.appendChild(sec);
  }});
}}

// ============================================================
// VIEW 4: COMBINED
// ============================================================
function renderCombined(area) {{
  const locs = Object.keys(SCHEDULE_DATA.locations||{{}});
  const locSlots = {{}};
  locs.forEach(loc=>{{
    let slots;
    if (_iter!=="Main"&&SCHEDULE_DATA.iterations&&SCHEDULE_DATA.iterations[_iter]&&SCHEDULE_DATA.iterations[_iter][loc]) {{
      slots = SCHEDULE_DATA.iterations[_iter][loc]||[];
    }} else {{
      slots = SCHEDULE_DATA.locations[loc]||[];
    }}
    locSlots[loc] = filterSlots(slots);
  }});

  // collect all times across all locations & days
  const allTimes = new Set();
  locs.forEach(loc=>locSlots[loc].forEach(s=>allTimes.add(s.time)));
  const times = [...allTimes].sort();

  // build lookup: loc -> day -> time -> slots
  const lookup={{}};
  locs.forEach(loc=>{{
    lookup[loc]={{}};
    locSlots[loc].forEach(s=>{{
      if (!lookup[loc][s.day_of_week]) lookup[loc][s.day_of_week]={{}};
      if (!lookup[loc][s.day_of_week][s.time]) lookup[loc][s.day_of_week][s.time]=[];
      lookup[loc][s.day_of_week][s.time].push(s);
    }});
  }});

  const LOC_ABBR = {{"Kwality House, Kemps Corner":"KW","Supreme HQ, Bandra":"SU","Kenkere House":"KE"}};
  const LOC_CLS  = {{"Kwality House, Kemps Corner":"ct-loc-kw","Supreme HQ, Bandra":"ct-loc-su","Kenkere House":"ct-loc-ke"}};

  const wrap = h("div",{{className:"combined-wrap"}});
  const table = h("table",{{className:"combined-table"}});
  const thead = h("thead");

  // header row: Time | KW Mon | SU Mon | KE Mon | gap | KW Tue | ...
  const hr1 = h("tr");
  hr1.appendChild(h("th",{{}}, "Time"));
  DAY_ORDER.forEach((day,di)=>{{
    locs.forEach(loc=>{{
      const abbr = LOC_ABBR[loc]||loc.slice(0,2);
      hr1.appendChild(h("th",{{}},`${{abbr}} ${{DAY_SHORT[di]}}`));
    }});
    if (di<6) hr1.appendChild(h("th",{{style:"background:var(--surface2);width:6px"}}));
  }});
  thead.appendChild(hr1);
  table.appendChild(thead);

  const tbody = h("tbody");
  times.forEach(t=>{{
    // only show row if at least 1 location has a class this time
    const hasAny = DAY_ORDER.some(day=>locs.some(loc=>(lookup[loc][day]&&lookup[loc][day][t]&&lookup[loc][day][t].length>0)));
    if (!hasAny) return;
    const prime=isPrime(t);
    const tr=h("tr");
    const timeCell = h("td",{{className:"ct-time"+(prime?" prime":"")}},t);
    tr.appendChild(timeCell);
    DAY_ORDER.forEach((day,di)=>{{
      locs.forEach(loc=>{{
        const td = h("td");
        const ss=(lookup[loc][day]&&lookup[loc][day][t])||[];
        ss.forEach(s=>{{
          const fam=getFamily(s.class_name);
          const card=h("div",{{
            className:"ct-card",
            style:`background:${{FAMILY_BG[fam]}};border-left:2px solid ${{FAMILY_BORDER[fam]}}`,
            onclick:()=>openModal(s)
          }});
          card.innerHTML=`<div style="font-size:10px;font-weight:700;color:${{FAMILY_TEXT[fam]}}">${{displayClass(s.class_name)}}</div><div style="font-size:9px;color:#64748B">${{s.trainer_1||""}}</div>`;
          td.appendChild(card);
        }});
        tr.appendChild(td);
      }});
      if (di<6) tr.appendChild(h("td",{{style:"background:var(--surface2)"}}));
    }});
    tbody.appendChild(tr);
  }});
  table.appendChild(tbody);
  wrap.appendChild(table);
  area.appendChild(wrap);
}}

// ============================================================
// VIEW 5: ANALYTICS
// ============================================================
function renderAnalytics(area, slots) {{
  const grid = h("div",{{className:"analytics-grid"}});

  // A: Class Mix
  grid.appendChild(makeClassMixCard(slots));

  // B: Fill Rate by Day + Band
  grid.appendChild(makeFillRateCard(slots));

  // C: Trainer Utilization
  grid.appendChild(makeTrainerUtilCard(slots));

  // D: Optimisation Opportunities
  grid.appendChild(makeOppCard(slots));

  // E: Score Distribution
  grid.appendChild(makeScoreDistCard(slots));

  area.appendChild(grid);
}}

function makeClassMixCard(slots) {{
  const card = h("div",{{className:"analytics-card full"}});
  card.appendChild(h("div",{{className:"analytics-title"}},"Class Mix"));

  const total = slots.length;
  if (total===0) {{ card.appendChild(h("div",{{}},"No data")); return card; }}

  const counts={{}};
  slots.forEach(s=>{{
    const fam=getFamily(s.class_name);
    counts[fam]=(counts[fam]||0)+1;
  }});

  const famOrder=["barre","powercycle","strength_lab","mat_57","recovery","foundations","hiit","default"];
  const LABELS={{barre:"Barre Family",powercycle:"PowerCycle",strength_lab:"Strength Lab",
    mat_57:"Mat 57",recovery:"Recovery",foundations:"Foundations",hiit:"HIIT",default:"Other"}};

  const svgW=700, svgH=40, barH=28, y=(svgH-barH)/2;
  let segments="";
  let x=0;
  famOrder.forEach(fam=>{{
    const cnt=counts[fam]||0;
    if (!cnt) return;
    const w=(cnt/total)*svgW;
    segments+=`<rect x="${{x.toFixed(1)}}" y="${{y}}" width="${{w.toFixed(1)}}" height="${{barH}}" fill="${{FAMILY_BORDER[fam]}}" rx="2"/>`;
    if (w>30) segments+=`<text x="${{(x+w/2).toFixed(1)}}" y="${{(y+barH/2+4).toFixed(1)}}" text-anchor="middle" font-size="11" font-weight="600" fill="#fff">${{(cnt/total*100).toFixed(0)}}%</text>`;
    x+=w;
  }});

  const svgEl=h("svg",{{viewBox:`0 0 ${{svgW}} ${{svgH}}`,style:"width:100%;height:auto;margin-bottom:10px;display:block"}});
  svgEl.innerHTML=segments;
  card.appendChild(svgEl);

  // legend
  const leg=h("div",{{className:"legend-row"}});
  famOrder.forEach(fam=>{{
    const cnt=counts[fam]||0;
    if (!cnt) return;
    leg.appendChild(h("div",{{className:"legend-item"}},[
      h("div",{{className:"legend-dot",style:`background:${{FAMILY_BORDER[fam]}}`}}),
      `${{LABELS[fam]}} (${{cnt}})`
    ]));
  }});
  card.appendChild(leg);
  return card;
}}

function makeFillRateCard(slots) {{
  const card = h("div",{{className:"analytics-card full"}});
  card.appendChild(h("div",{{className:"analytics-title"}},"Fill Rate by Day and Time Band"));

  // by day
  const byDay={{}};
  DAY_ORDER.forEach(d=>byDay[d]={{sum:0,cnt:0}});
  slots.forEach(s=>{{
    if (!byDay[s.day_of_week]) return;
    byDay[s.day_of_week].sum+=(s.predicted_fill_rate||0);
    byDay[s.day_of_week].cnt++;
  }});

  const svgW=640, svgH=100, barW=60, gap=20, padL=40, padT=10;
  let bars="", labels="";
  DAY_ORDER.forEach((d,i)=>{{
    const avg = byDay[d].cnt ? byDay[d].sum/byDay[d].cnt : 0;
    const bh = avg*(svgH-padT-20);
    const x=padL+i*(barW+gap);
    const clr=fillClr(avg);
    bars+=`<rect x="${{x}}" y="${{svgH-20-bh.toFixed(1)}}" width="${{barW}}" height="${{bh.toFixed(1)}}" fill="${{clr}}" rx="3"/>`;
    labels+=`<text x="${{x+barW/2}}" y="${{svgH-4}}" text-anchor="middle" font-size="11" fill="#64748B">${{DAY_SHORT[i]}}</text>`;
    if (byDay[d].cnt) bars+=`<text x="${{x+barW/2}}" y="${{(svgH-20-bh-5).toFixed(1)}}" text-anchor="middle" font-size="10" font-weight="600" fill="${{clr}}">${{pct(avg)}}</text>`;
  }});

  const svg=h("svg",{{viewBox:`0 0 ${{svgW}} ${{svgH}}`,style:"width:100%;height:auto;margin-bottom:16px;display:block"}});
  svg.innerHTML=bars+labels;
  card.appendChild(svg);

  // by band
  card.appendChild(h("div",{{className:"analytics-sub"}},"By time band"));
  const bands={{morning:{{sum:0,cnt:0}},midday:{{sum:0,cnt:0}},evening:{{sum:0,cnt:0}}}};
  slots.forEach(s=>{{
    const b=timeBand(s.time);
    if (bands[b]) {{ bands[b].sum+=(s.predicted_fill_rate||0); bands[b].cnt++; }}
  }});
  Object.entries(bands).forEach(([band,{{sum,cnt}}])=>{{
    const avg=cnt?sum/cnt:0;
    const row=h("div",{{style:"display:flex;align-items:center;gap:10px;margin-bottom:6px"}});
    row.appendChild(h("span",{{style:"font-size:12px;font-weight:600;width:70px;text-transform:capitalize;color:var(--text-muted)"}},band));
    const barBg=h("div",{{className:"util-bar-bg",style:"flex:1"}});
    const barFill=h("div",{{className:"util-bar-fill",style:`width:${{(avg*100).toFixed(0)}}%;background:${{fillClr(avg)}}`}});
    barBg.appendChild(barFill);
    row.appendChild(barBg);
    row.appendChild(h("span",{{style:`font-size:12px;font-weight:700;width:38px;text-align:right;color:${{fillClr(avg)}}`}},pct(avg)));
    card.appendChild(row);
  }});
  return card;
}}

function makeTrainerUtilCard(slots) {{
  const card=h("div",{{className:"analytics-card"}});
  card.appendChild(h("div",{{className:"analytics-title"}},"Trainer Utilization"));

  const byT={{}};
  slots.forEach(s=>{{
    const t=s.trainer_1||"?";
    if (!byT[t]) byT[t]={{cnt:0,hours:0,fillSum:0}};
    byT[t].cnt++;
    byT[t].hours+=(s.duration_min||57)/60;
    byT[t].fillSum+=(s.predicted_fill_rate||0);
  }});

  const rows=Object.entries(byT).sort((a,b)=>b[1].cnt-a[1].cnt);

  const hdr=h("div",{{className:"trainer-util-row"}});
  ["Trainer","Cls","Hrs","Avg Fill","Utilization"].forEach(l=>{{
    hdr.appendChild(h("div",{{style:"font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.07em;color:var(--text-muted)"}},l));
  }});
  card.appendChild(hdr);

  rows.forEach(([trainer,{{cnt,hours,fillSum}}])=>{{
    const avg=cnt?fillSum/cnt:0;
    const util=Math.min(1,hours/15);
    const row=h("div",{{className:"trainer-util-row"}});
    row.appendChild(h("div",{{style:"font-size:12px;font-weight:600;color:var(--text);overflow:hidden;text-overflow:ellipsis;white-space:nowrap"}},trainer));
    row.appendChild(h("div",{{style:"font-size:12px;font-weight:700;text-align:center;color:var(--primary)"}},String(cnt)));
    row.appendChild(h("div",{{style:"font-size:12px;text-align:center;color:var(--text-muted)"}},hours.toFixed(1)));
    row.appendChild(h("div",{{style:`font-size:12px;font-weight:700;text-align:center;color:${{fillClr(avg)}}`}},pct(avg)));
    const barBg=h("div",{{className:"util-bar-bg"}});
    const barClr=util>=0.8?"#16A34A":util>=0.5?"#D97706":"#DC2626";
    const barFill=h("div",{{className:"util-bar-fill",style:`width:${{(util*100).toFixed(0)}}%;background:${{barClr}}`}});
    barBg.appendChild(barFill);
    row.appendChild(barBg);
    card.appendChild(row);
  }});
  return card;
}}

function makeOppCard(slots) {{
  const card=h("div",{{className:"analytics-card"}});
  card.appendChild(h("div",{{className:"analytics-title"}},"Optimisation Opportunities"));

  OPTIMISATION_OPPORTUNITIES.forEach(opp=>{{
    const oc=h("div",{{className:"opp-card"}});
    oc.appendChild(h("div",{{className:"opp-text"}},opp));
    card.appendChild(oc);
  }});

  // violation alerts
  const viols=[];
  slots.forEach(s=>{{ (s.constraint_violations||[]).forEach(v=>viols.push({{v,s}})); }});
  if (viols.length>0) {{
    card.appendChild(h("div",{{style:"margin-top:12px;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:var(--danger)"}},`${{viols.length}} Constraint Violations`));
    viols.slice(0,6).forEach(({{}})=>{{
      const vc=h("div",{{className:"opp-card red"}});
      vc.appendChild(h("div",{{className:"opp-text"}},`${{viols[0].s.day_of_week}} ${{viols[0].s.time}}: ${{viols[0].v}}`));
      card.appendChild(vc);
    }});
  }}
  return card;
}}

function makeScoreDistCard(slots) {{
  const card=h("div",{{className:"analytics-card full"}});
  card.appendChild(h("div",{{className:"analytics-title"}},"Score Distribution (0–100)"));

  const buckets=Array(10).fill(0);
  slots.forEach(s=>{{
    const sc=Math.min(99,Math.max(0,s.score||0));
    buckets[Math.floor(sc/10)]++;
  }});
  const maxB=Math.max(...buckets,1);
  const svgW=620, svgH=80, bW=42, gap=20, padL=30, padT=10;
  let rects="",xlbls="",ylbl="";
  buckets.forEach((cnt,i)=>{{
    const bh=(cnt/maxB)*(svgH-padT-20);
    const x=padL+i*(bW+gap);
    const clr=i>=7?"#16A34A":i>=4?"#D97706":"#DC2626";
    rects+=`<rect x="${{x}}" y="${{svgH-20-bh.toFixed(1)}}" width="${{bW}}" height="${{bh.toFixed(1)}}" fill="${{clr}}" rx="3"/>`;
    xlbls+=`<text x="${{x+bW/2}}" y="${{svgH-4}}" text-anchor="middle" font-size="10" fill="#64748B">${{i*10}}-${{i*10+9}}</text>`;
    if (cnt) rects+=`<text x="${{x+bW/2}}" y="${{(svgH-20-bh-4).toFixed(1)}}" text-anchor="middle" font-size="10" font-weight="600" fill="${{clr}}">${{cnt}}</text>`;
  }});
  const svg=h("svg",{{viewBox:`0 0 ${{svgW}} ${{svgH}}`,style:"width:100%;height:auto;display:block"}});
  svg.innerHTML=rects+xlbls+ylbl;
  card.appendChild(svg);
  return card;
}}

// ============================================================
// MODAL
// ============================================================
function openModal(slot) {{
  const fam=getFamily(slot.class_name);
  const famClr=FAMILY_BORDER[fam]||"#1E40AF";
  const box=document.getElementById("modal-box");

  // find trainer's other classes this week from current filtered slots
  const allSlots=getSlots();
  const trainerWeek=allSlots.filter(s=>s.trainer_1===slot.trainer_1&&!(s.day_of_week===slot.day_of_week&&s.time===slot.time&&s.class_name===slot.class_name));

  const viols=slot.constraint_violations||[];
  const violsHtml=viols.length
    ? viols.map(v=>`<span class="viol-tag">${{v}}</span>`).join("")
    : '<span style="color:var(--success);font-size:12px;font-weight:600">&#10003; No violations</span>';

  const lcRate=(slot.historical_late_cancel_rate||0);
  const nsRate=(slot.historical_no_show_rate||0);

  const twHtml=trainerWeek.slice(0,8).map(s=>
    `<div class="tw-chip"><strong>${{s.day_of_week.slice(0,3)}}</strong> ${{s.time}} · ${{displayClass(s.class_name)}}</div>`
  ).join("");

  box.innerHTML=`
    <div class="modal-hdr">
      <div style="flex:1">
        <div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:var(--text-muted);margin-bottom:6px">
          ${{slot.day_of_week}} &middot; ${{slot.time}} &middot; ${{displayRoom(slot.room)||"Studio"}} &middot; ${{slot.duration_min||57}} min
        </div>
        <div style="font-size:20px;font-weight:800;color:${{famClr}};margin-bottom:4px">${{displayClass(slot.class_name)}}</div>
        <div style="font-size:13px;color:var(--text-muted);margin-bottom:8px">
          ${{slot.trainer_1||"—"}} &middot; Capacity: ${{slot.capacity||"—"}}
        </div>
        <div class="chip-row">
          <span class="chip cbadge-${{slot.recommendation||"INCLUDE"}}">${{slot.recommendation||"INCLUDE"}}</span>
          ${{slot.is_experimental?'<span class="chip chip-exp">EXPERIMENTAL</span>':''}}
          ${{isPrime(slot.time)?'<span class="chip chip-prime">&#11088; Prime Slot</span>':''}}
        </div>
      </div>
      <button class="modal-close-btn" onclick="closeModal()">&#10005;</button>
    </div>
    <div class="modal-body">
      <div class="modal-section">
        <div class="modal-section-ttl">Scheduling Reason</div>
        <div class="modal-reason-box">${{slot.scheduling_reason||"—"}}</div>
      </div>
      <div class="modal-section">
        <div class="modal-section-ttl">Performance Metrics</div>
        <div class="modal-metric-grid">
          <div class="modal-metric">
            <div class="mm-lbl">Predicted Fill Rate</div>
            <div class="mm-val" style="color:${{fillClr(slot.predicted_fill_rate)}}">${{pct(slot.predicted_fill_rate)}}</div>
          </div>
          <div class="modal-metric">
            <div class="mm-lbl">Historical Fill Rate</div>
            <div class="mm-val" style="color:${{fillClr(slot.historical_avg_fill)}}">${{pct(slot.historical_avg_fill)}}</div>
          </div>
          <div class="modal-metric">
            <div class="mm-lbl">Avg Check-in</div>
            <div class="mm-val" style="color:var(--primary)">${{(slot.historical_avg_checkin||0).toFixed(1)}}</div>
          </div>
          <div class="modal-metric">
            <div class="mm-lbl">Historical Sessions</div>
            <div class="mm-val" style="color:var(--text)">${{slot.historical_session_count||0}}</div>
          </div>
          <div class="modal-metric">
            <div class="mm-lbl">Optimisation Score</div>
            <div class="mm-val" style="color:var(--accent)">${{(slot.score||0).toFixed(1)}}</div>
          </div>
          <div class="modal-metric">
            <div class="mm-lbl">Trainer Overall Fill</div>
            <div class="mm-val" style="color:${{fillClr(slot.trainer_overall_fill)}}">${{pct(slot.trainer_overall_fill||0)}}</div>
          </div>
        </div>
      </div>
      <div class="modal-section">
        <div class="modal-section-ttl">Trainer Profile</div>
        <div class="modal-metric-grid">
          <div class="modal-metric">
            <div class="mm-lbl">Sessions at Location</div>
            <div class="mm-val" style="color:var(--primary)">${{slot.trainer_total_sessions||"—"}}</div>
          </div>
          <div class="modal-metric">
            <div class="mm-lbl">Trainer Avg Check-in</div>
            <div class="mm-val" style="color:var(--success)">${{(slot.trainer_overall_checkin||0).toFixed(1)}}</div>
          </div>
        </div>
      </div>
      <div class="modal-section">
        <div class="modal-section-ttl">Late Cancel &amp; No-Show</div>
        <div style="margin-bottom:8px">
          <div style="font-size:11px;color:var(--text-muted);margin-bottom:3px">Late Cancel Rate: <strong style="color:${{lcRate>0.15?"var(--danger)":"var(--success)}}">${{pct(lcRate)}}</strong></div>
          <div class="prog-bar"><div class="prog-fill" style="width:${{(lcRate*100).toFixed(0)}}%;background:${{lcRate>0.15?"var(--danger)":"var(--success)"}}"></div></div>
        </div>
        <div>
          <div style="font-size:11px;color:var(--text-muted);margin-bottom:3px">No-Show Rate: <strong style="color:${{nsRate>0.15?"var(--danger)":"var(--success)}}">${{pct(nsRate)}}</strong></div>
          <div class="prog-bar"><div class="prog-fill" style="width:${{(nsRate*100).toFixed(0)}}%;background:${{nsRate>0.15?"var(--danger)":"var(--success)"}}"></div></div>
        </div>
      </div>
      <div class="modal-section">
        <div class="modal-section-ttl">Constraint Status</div>
        <div>${{violsHtml}}</div>
      </div>
      ${{trainerWeek.length?`
      <div class="modal-section">
        <div class="modal-section-ttl">This trainer's other classes this week</div>
        <div class="trainer-week-row">${{twHtml}}</div>
      </div>`:"" }}
    </div>
  `;
  document.getElementById("modal-overlay").classList.add("open");
}}

function closeModal() {{
  document.getElementById("modal-overlay").classList.remove("open");
}}

document.getElementById("modal-overlay").addEventListener("click", e=>{{
  if (e.target===document.getElementById("modal-overlay")) closeModal();
}});

document.addEventListener("keydown", e=>{{
  if (e.key==="Escape") closeModal();
}});
</script>

{_rules_panel_html()}
</body>
</html>"""

        (WEB_DIR / "index.html").write_text(html)
        print(f"[Agent 6] Web interface written to {WEB_DIR}/index.html")
        print(f"[Agent 6] To view with rule toggles, run:")
        print(f"          python3 serve.py --week {self._week_label} --port 8080")

    # ------------------------------------------------------------------ #
    #  Rich summary + assertions
    # ------------------------------------------------------------------ #

  def _print_summary(self, scorecard):
        console = Console()
        console.print("\n[bold blue]Studio Schedule — Summary[/bold blue]\n")
        for loc, entry in scorecard["locations"].items():
            console.print(f"[bold green]{loc}[/bold green]")
            t = Table(box=box.SIMPLE)
            t.add_column("Metric", style="cyan")
            t.add_column("Value")
            t.add_row("Total Classes", str(entry.get("total_classes", 0)))
            t.add_row("Predicted Avg Fill", f"{entry.get('predicted_avg_fill_rate', 0):.1%}")
            t.add_row("Experimental %", f"{entry.get('experimental_pct', 0):.0%}")
            console.print(t)
            if entry.get("hard_constraint_violations"):
                console.print("[bold red]  VIOLATIONS:[/bold red]")
                for v in entry["hard_constraint_violations"]:
                    console.print(f"  [red]x {v}[/red]")
            console.print("[bold yellow]  Top 3 Opportunities:[/bold yellow]")
            for opp in entry.get("optimisation_opportunities", [])[:3]:
                console.print(f"  [yellow]-> {opp}[/yellow]")
            console.print()

  def _run_assertions(self, scorecard, by_location):
        errors = []
        warnings = []
        config_errors, config_warnings = self._validate_against_schedule_config(by_location)
        errors.extend(config_errors)
        warnings.extend(config_warnings)
        kw = scorecard["locations"].get("Kwality House, Kemps Corner", {})
        ke_slots = by_location.get("Kenkere House", [])
        su_slots = by_location.get("Supreme HQ, Bandra", [])
        kw_slots = by_location.get("Kwality House, Kemps Corner", [])

        expected_totals = {
            "Kwality House, Kemps Corner": (60, 90),
            "Supreme HQ, Bandra": (55, 85),
            "Kenkere House": (45, 70),
        }
        for loc, (lo, hi) in expected_totals.items():
            actual = scorecard["locations"].get(loc, {}).get("total_classes", 0)
            if actual < lo or actual > hi:
                errors.append(f"{loc}: total classes {actual} outside [{lo},{hi}]")
        for slot in ke_slots:
            if "PowerCycle" in slot.get("class_name", ""):
                errors.append(f"PowerCycle at Kenkere: {slot['time']} {slot['day_of_week']}")
                break
        for slot in ke_slots + su_slots:
            if "Strength Lab" in slot.get("class_name", ""):
                errors.append(f"Strength Lab at {slot['location']}")
                break
        for slot in kw_slots + su_slots + ke_slots:
            if slot.get("day_of_week") == "Sunday" and int(slot["time"][:2]) < 10:
                errors.append(f"Sunday class before 10:00 at {slot['location']}")
        for slot in kw_slots + su_slots + ke_slots:
            if not slot.get("trainer_1"):
                errors.append(f"Empty trainer_1 at {slot['location']} {slot['day_of_week']} {slot['time']}")
            if "Unknown Class" in slot.get("class_name", ""):
                errors.append(f"Unknown Class scheduled at {slot['location']} {slot['day_of_week']} {slot['time']}")

        trainer_day_shifts = defaultdict(set)
        for slot in kw_slots + su_slots + ke_slots:
            trainer = slot.get("trainer_1")
            if not trainer:
                continue
            shift = "AM" if int(slot["time"][:2]) < 13 else "PM"
            trainer_day_shifts[(trainer, slot.get("date") or slot["day_of_week"])].add(shift)
        for (trainer, day), shifts in trainer_day_shifts.items():
            if len(shifts) > 1:
                errors.append(f"{trainer} teaches both AM and PM on {day}")

        trainer_weekly_minutes = defaultdict(int)
        for slot in kw_slots + su_slots + ke_slots:
            trainer = slot.get("trainer_1")
            if trainer:
                trainer_weekly_minutes[trainer] += int(slot.get("duration_min") or 57)
        for trainer, minutes in trainer_weekly_minutes.items():
            if minutes > 15 * 60:
                errors.append(f"{trainer} weekly load {minutes / 60:.1f}h exceeds 15.0h cap")

        # New: Recovery must be last in its shift on each (loc, day)
        from collections import defaultdict as _dd
        def _shift(time_str):
            h = int(time_str[:2])
            if 7 <= h < 12: return "morning"
            if 12 <= h < 16: return "midday"
            return "evening"

        per_day = _dd(list)
        for slot in kw_slots + su_slots + ke_slots:
            per_day[(slot["location"], slot["day_of_week"])].append(slot)
        for (loc, day), day_slots in per_day.items():
            for s in day_slots:
                if "Recovery" not in s["class_name"]:
                    continue
                rec_shift = _shift(s["time"])
                rec_min = int(s["time"][:2]) * 60 + int(s["time"][3:5])
                for other in day_slots:
                    if other is s: continue
                    if _shift(other["time"]) != rec_shift: continue
                    om = int(other["time"][:2]) * 60 + int(other["time"][3:5])
                    if om > rec_min:
                        errors.append(f"Recovery not last in shift: {loc} {day} {s['time']} (later: {other['time']} {other['class_name']})")

        # Per-location format limits
        schedule_config = {}
        config_path = Path("config/schedule_config.json")
        if config_path.exists():
            with open(config_path) as f:
                schedule_config = json.load(f)

        def _mix_limits(loc: str, class_name: str, default_min=None, default_max=None):
            entry = (
                schedule_config
                .get("class_mix", {})
                .get(loc, {})
                .get(class_name, {})
            )
            return entry.get("min", default_min), entry.get("max", default_max)

        for loc, entry in scorecard["locations"].items():
            if entry.get("schedule_score", 0) < MIN_SCHEDULE_SCORE:
                errors.append(
                    f"{loc}: schedule score {entry.get('schedule_score', 0):.1f}/100 < {MIN_SCHEDULE_SCORE:.0f}/100 target"
                )
            fc = entry.get("format_counts", {})
            _, hiit_max = _mix_limits(loc, "Studio HIIT", None, 3)
            if hiit_max is not None and fc.get("Studio HIIT", 0) > hiit_max:
                warnings.append(f"{loc}: HIIT count {fc['Studio HIIT']} above soft target {hiit_max}/week")
            _, amped_max = _mix_limits(loc, "Studio Amped Up!", None, 2)
            if amped_max is not None and fc.get("Studio Amped Up!", 0) > amped_max:
                warnings.append(f"{loc}: Amped Up count {fc['Studio Amped Up!']} above soft target {amped_max}/week")
            _, back_body_max = _mix_limits(loc, "Studio Back Body Blaze", None, 3)
            if back_body_max is not None and fc.get("Studio Back Body Blaze", 0) > back_body_max:
                warnings.append(f"{loc}: Back Body Blaze count {fc['Studio Back Body Blaze']} above soft target {back_body_max}/week")

            # Barre family pct >= 25%
            if entry.get("barre_family_pct", 0) < 0.25:
                warnings.append(f"{loc}: Barre family pct {entry['barre_family_pct']:.1%} < 25%")

            # Format floors per location
            if loc == "Kenkere House":
                pc = fc.get("Studio PowerCycle", 0)
                if pc != 0:
                    errors.append(f"Kenkere PowerCycle count = {pc}, must be 0")
            else:
                # Family floors
                if entry.get("barre_family_count", 0) < 14:
                    warnings.append(f"{loc}: Barre family count {entry.get('barre_family_count', 0)} < 14 floor")
                mat_min, _ = _mix_limits(loc, "Studio Mat 57", 4, None)
                if mat_min and fc.get("Studio Mat 57", 0) < mat_min:
                    warnings.append(f"{loc}: Mat 57 count {fc.get('Studio Mat 57', 0)} < {mat_min} floor")
                cardio_min, _ = _mix_limits(loc, "Studio Cardio Barre", 4, None)
                if cardio_min and fc.get("Studio Cardio Barre", 0) < cardio_min:
                    warnings.append(f"{loc}: Cardio Barre count {fc.get('Studio Cardio Barre', 0)} < {cardio_min} floor")
                fit_min, _ = _mix_limits(loc, "Studio FIT", 5, None)
                if fit_min and fc.get("Studio FIT", 0) < fit_min:
                    warnings.append(f"{loc}: FIT count {fc.get('Studio FIT', 0)} < {fit_min} floor")
                if loc == "Kwality House, Kemps Corner":
                    sl = fc.get("Studio Strength Lab", 0)
                    sl_min, sl_max = _mix_limits(loc, "Studio Strength Lab", 6, 8)
                    if sl < sl_min or (sl_max is not None and sl > sl_max):
                        warnings.append(f"Kwality Strength Lab count {sl} outside [{sl_min},{sl_max}]")
                    pc_min, _ = _mix_limits(loc, "Studio PowerCycle", 6, None)
                    if pc_min and fc.get("Studio PowerCycle", 0) < pc_min:
                        warnings.append(f"Kwality PowerCycle count {fc.get('Studio PowerCycle', 0)} < {pc_min} floor")
                if loc == "Supreme HQ, Bandra":
                    pc_min, _ = _mix_limits(loc, "Studio PowerCycle", 14, None)
                    if pc_min and fc.get("Studio PowerCycle", 0) < pc_min:
                        warnings.append(f"Supreme PowerCycle count {fc.get('Studio PowerCycle', 0)} < {pc_min} floor")

        if warnings:
            print("[OUTPUT QUALITY WARNINGS]")
            for w in warnings:
                print(f"  WARN: {w}")
        if errors:
            print("[OUTPUT QUALITY FAILURES]")
            for e in errors:
                print(f"  ERROR: {e}")
            raise AssertionError(f"{len(errors)} output quality failure(s)")
        print("[Output quality checks passed]")

  def _run_iteration_score_assertions(self, all_by_location, all_schedules, iteration_names):
        errors = []
        for idx, loc_map in enumerate(all_by_location[:3]):
            schedule = all_schedules[idx].get("schedule", []) if idx < len(all_schedules) else []
            baselines = self._build_score_baselines(schedule)
            iteration_name = iteration_names[idx] if idx < len(iteration_names) else f"Iteration {idx + 1}"
            trainer_weekly_minutes = defaultdict(int)
            for slot in schedule:
                trainer = slot.get("trainer_1")
                if trainer:
                    trainer_weekly_minutes[trainer] += int(slot.get("duration_min") or 57)
            for trainer, minutes in trainer_weekly_minutes.items():
                if minutes > 15 * 60:
                    errors.append(f"{iteration_name}: {trainer} weekly load {minutes / 60:.1f}h exceeds 15.0h cap")
            for loc, slots in loc_map.items():
                entry = self._build_scorecard_entry(loc, slots, baselines)
                if entry.get("schedule_score", 0) < MIN_SCHEDULE_SCORE:
                    errors.append(
                        f"{iteration_name} / {loc}: schedule score "
                        f"{entry.get('schedule_score', 0):.1f}/100 < {MIN_SCHEDULE_SCORE:.0f}/100 target"
                    )
        if errors:
            print("[ITERATION SCORE FAILURES]")
            for e in errors:
                print(f"  ERROR: {e}")
            raise AssertionError(f"{len(errors)} iteration score failure(s)")
        print("[Iteration score checks passed]")
